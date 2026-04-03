from flask import Flask, render_template, request, redirect, url_for, session, send_from_directory, jsonify, flash, Response
# NOTE: Reklamacje MVP (PostgreSQL)
from functools import wraps
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
import datetime
import csv
import io
import json
import uuid
import re
import shutil
import hashlib
import time
import subprocess
from contextlib import closing
from pathlib import Path
import requests
import sqlite3

app = Flask(__name__, static_folder='static', static_url_path='/static')
app.secret_key = os.environ.get('FLASK_SECRET_KEY', 'suuhouse-magazyn-secret-key-change-in-production-2025')
app.config['UPLOAD_FOLDER'] = 'shared'
app.config['PERMANENT_SESSION_LIFETIME'] = datetime.timedelta(minutes=60)
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
if os.environ.get('FLASK_ENV') == 'production' or os.environ.get('HTTPS', '').lower() in ('1', 'true', 'on'):
    app.config['SESSION_COOKIE_SECURE'] = True
app.config['REKLAMACJE_UPLOADS'] = os.environ.get(
    "REKLAMACJE_UPLOADS",
    os.path.join(os.path.dirname(__file__), "uploads", "reklamacje"),
)
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size

# Dozwolone rozszerzenia plików
ALLOWED_EXTENSIONS = {
    'txt', 'pdf', 'png', 'jpg', 'jpeg', 'gif', 'webp',
    'doc', 'docx', 'xls', 'xlsx', 'zip', 'rar',
    'mp4', 'avi', 'mov', 'mp3'
}

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def compress_uploaded_image_in_place(file_path: str) -> bool:
    """
    Agresywna kompresja obrazów po uploadzie.
    - JPG/JPEG: strip EXIF + resize max 1920px + quality ~82 + jpegoptim.
    - PNG: tylko strip + resize (bez konwersji do JPG).
    Funkcja jest best-effort: w razie błędu nie przerywa uploadu.
    """
    try:
        path = str(file_path or "")
        if not path or not os.path.exists(path):
            return False
        ext = os.path.splitext(path)[1].lower().lstrip(".")
        if ext not in {"jpg", "jpeg", "png"}:
            return False

        mogrify = shutil.which("mogrify")
        if not mogrify:
            return False

        # Resize/strip (dla JPG/JPEG dodatkowo quality)
        cmd = [mogrify, "-strip", "-resize", "1920x1920>"]
        if ext in {"jpg", "jpeg"}:
            cmd += ["-quality", "82"]
        cmd.append(path)
        subprocess.run(cmd, check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=45)

        if ext in {"jpg", "jpeg"}:
            jpegoptim = shutil.which("jpegoptim")
            if jpegoptim:
                subprocess.run([jpegoptim, "-q", "--strip-all", "--max=82", path], check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=45)
        return True
    except Exception:
        return False


def _optimize_kontakt_logo_in_place(path: str) -> bool:
    try:
        p = str(path or "")
        if not p or not os.path.exists(p):
            return False
        ext = os.path.splitext(p)[1].lower().lstrip(".")
        if ext not in {"jpg", "jpeg", "png", "gif", "webp"}:
            return False
        mogrify = shutil.which("mogrify")
        if not mogrify:
            return False
        cmd = [mogrify, "-strip", "-resize", "512x512>"]
        if ext in {"jpg", "jpeg"}:
            cmd += ["-quality", "85"]
        elif ext == "webp":
            cmd += ["-quality", "82"]
        cmd.append(p)
        subprocess.run(cmd, check=False, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL, timeout=30)
        if ext in {"jpg", "jpeg"}:
            jpegoptim = shutil.which("jpegoptim")
            if jpegoptim:
                subprocess.run(
                    [jpegoptim, "-q", "--strip-all", "--max=85", p],
                    check=False,
                    stdout=subprocess.DEVNULL,
                    stderr=subprocess.DEVNULL,
                    timeout=30,
                )
        return True
    except Exception:
        return False


# Google Sheets / Drive file (stan magazynu) – fallback
GOOGLE_SHEET_ID = '1qN8sUUUXv1PXjjVLoEwhCeoTI5XDUQzC4gHQLNkhvhg'
GOOGLE_SHEET_GID = '0'

# n8n webhook (nowe źródło magazynu)
N8N_INVENTORY_URL = os.environ.get("N8N_INVENTORY_URL", "").strip()
N8N_INVENTORY_API_KEY = os.environ.get("N8N_INVENTORY_API_KEY", "").strip()

CACHE_BUSTER = "20260108-21"
REKLAMACJE_DB_URL = os.environ.get("REKLAMACJE_DB_URL", "").strip()
REKLAMACJE_DB_PATH = os.environ.get(
    "REKLAMACJE_DB_PATH",
    os.path.join(os.path.dirname(__file__), "reklamacje.db"),
)
REKLAMACJE_JSON_FILE = os.path.join(os.path.dirname(__file__), "reklamacje.json")
KARTONY_JSON_FILE = os.path.join(os.path.dirname(__file__), "kartony.json")
PLANER_JSON_FILE = os.path.join(os.path.dirname(__file__), "planer.json")
KONTAKTY_JSON_FILE = os.path.join(os.path.dirname(__file__), "shared", "kontakty.json")
KONTAKTY_LOGO_DIR = os.path.join(os.path.dirname(__file__), "static", "kontakty")
KONTAKTY_LOGO_URL_PREFIX = "/static/kontakty/"
KONTAKTY_LOGO_EXT = {".png", ".jpg", ".jpeg", ".gif", ".webp"}
SKU_INDEX_CACHE_FILE = os.path.join(os.path.dirname(__file__), "sku_index_cache.json")
SKU_MANUAL_FILE = os.path.join(os.path.dirname(__file__), "sku_manual.json")
REKLAMACJE_STATUSES = [
    "Nowa",
    "W trakcie",
    "Czekamy na informację",
    "Odrzucona",
    "Zamknięta",
]
REKLAMACJE_TYPES = [
    "Uszkodzony element",
    "Uszkodzony cały produkt",
    "Brak elementu",
    "Zgłoszenie niereklamacyjne",
]
REKLAMACJE_DECISIONS = [
    "Dosłanie części",
    "Wymiana produktu",
    "Zwrot środków",
    "Odrzucenie",
]

EMAIL_PATTERN = re.compile(r"^[^@\s]+@[^@\s]+\.[a-zA-Z]{2,}$", re.IGNORECASE)


def validate_email(value):
    return bool(EMAIL_PATTERN.match(value or ""))


def validate_phone(value):
    return (value or "").isdigit()

def default_kartony_data():
    # Minimalny start — można edytować z UI na /kartony
    return {
        "products": [
            {
                "name": "Pergola Skyline 3x4",
                "tomek": False,
                "pdf_file": "",
                "has_instruction": False,
                "boxes": [
                    {"name": "Karton 1/4", "parts": ["Profil A", "Profil B", "Śruby M6 (zestaw)"]},
                    {"name": "Karton 2/4", "parts": ["Noga lewa", "Noga prawa", "Zaślepki"]},
                    {"name": "Karton 3/4", "parts": ["Belka poprzeczna", "Łączniki", "Instrukcja"]},
                    {"name": "Karton 4/4", "parts": ["Poszycie", "Akcesoria montażowe"]},
                ],
            }
        ]
    }

def load_kartony_data():
    try:
        if not os.path.exists(KARTONY_JSON_FILE):
            data = default_kartony_data()
            with open(KARTONY_JSON_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return data
        with open(KARTONY_JSON_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and isinstance(data.get("products"), list):
                return data
    except Exception:
        pass
    return default_kartony_data()

def save_kartony_data(data):
    with open(KARTONY_JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def default_planer_data():
    return {"rows": []}

def load_planer_data():
    try:
        if not os.path.exists(PLANER_JSON_FILE):
            data = default_planer_data()
            with open(PLANER_JSON_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return data
        with open(PLANER_JSON_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and isinstance(data.get("rows"), list):
                return data
    except Exception:
        pass
    return default_planer_data()

def save_planer_data(data):
    with open(PLANER_JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _ensure_kontakty_logo_dir():
    try:
        os.makedirs(KONTAKTY_LOGO_DIR, exist_ok=True)
    except Exception:
        pass


def default_kontakty_data():
    return {"contacts": []}


def load_kontakty_data():
    try:
        if not os.path.exists(KONTAKTY_JSON_FILE):
            data = default_kontakty_data()
            with open(KONTAKTY_JSON_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return data
        with open(KONTAKTY_JSON_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and isinstance(data.get("contacts"), list):
                return data
    except Exception:
        pass
    return default_kontakty_data()


def save_kontakty_data(data):
    with open(KONTAKTY_JSON_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def _kontakty_logo_disk_path(logo_url):
    if not logo_url or not isinstance(logo_url, str):
        return None
    u = logo_url.strip()
    if not u.startswith(KONTAKTY_LOGO_URL_PREFIX):
        return None
    name = u.rsplit("/", 1)[-1]
    if not name or ".." in name or "/" in name or "\\" in name:
        return None
    base = os.path.realpath(KONTAKTY_LOGO_DIR)
    fp = os.path.realpath(os.path.join(KONTAKTY_LOGO_DIR, name))
    if not fp.startswith(base + os.sep) and fp != base:
        return None
    return fp if os.path.isfile(fp) else None


def _delete_kontakt_logo_file(logo_url):
    p = _kontakty_logo_disk_path(logo_url)
    if p:
        try:
            os.remove(p)
        except Exception:
            pass


def _save_kontakt_logo_file(file_storage, contact_id):
    if not file_storage or not getattr(file_storage, "filename", None):
        return None
    ext = os.path.splitext(file_storage.filename)[1].lower()
    if ext not in KONTAKTY_LOGO_EXT:
        return None
    _ensure_kontakty_logo_dir()
    fname = f"{contact_id}{ext}"
    path = os.path.join(KONTAKTY_LOGO_DIR, fname)
    file_storage.save(path)
    _optimize_kontakt_logo_in_place(path)
    return KONTAKTY_LOGO_URL_PREFIX + fname


def _normalize_kontakt_fields(payload):
    if not isinstance(payload, dict):
        payload = {}
    return {
        "full_name": str(payload.get("full_name") or "").strip()[:200],
        "company": str(payload.get("company") or "").strip()[:200],
        "phone": str(payload.get("phone") or "").strip()[:80],
        "email": str(payload.get("email") or "").strip()[:200],
    }


def normalize_sku(value):
    txt = str(value or "").strip().upper()
    txt = re.sub(r"\s+", " ", txt)
    if not txt:
        return ""

    # Jeśli komórka ma format "SKU:XXXX", traktujemy SKU jako część po ":".
    m = re.search(r"\bSKU\s*:\s*([A-Z0-9._\-/]+)", txt)
    if m:
        txt = m.group(1)
    elif txt.startswith("SKU:"):
        txt = txt.split(":", 1)[1].strip()

    return txt.strip(" ,;|")

def _unique_skus(values):
    out = []
    seen = set()
    for v in values:
        sku = normalize_sku(v)
        if not sku or sku in seen:
            continue
        seen.add(sku)
        out.append(sku)
    return out

def default_sku_index_cache():
    return {"updated_at": "", "items": [], "suppliers": {}, "warning": {"active": False}, "snapshot_at": ""}

def default_sku_manual():
    return {"items": []}

def load_sku_index_cache():
    try:
        if os.path.exists(SKU_INDEX_CACHE_FILE) and os.path.getsize(SKU_INDEX_CACHE_FILE) > 2:
            with open(SKU_INDEX_CACHE_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict) and isinstance(data.get("items"), list):
                    data["items"] = _unique_skus(data.get("items"))
                    suppliers = data.get("suppliers")
                    if not isinstance(suppliers, dict):
                        suppliers = {}
                    data["suppliers"] = {normalize_sku(k): str(v or "").strip() for k, v in suppliers.items() if normalize_sku(k)}
                    warning = data.get("warning")
                    data["warning"] = warning if isinstance(warning, dict) else {"active": False}
                    data["snapshot_at"] = str(data.get("snapshot_at") or "")
                    return data
    except Exception:
        pass
    return default_sku_index_cache()

def save_sku_index_cache(data):
    raw_suppliers = (data or {}).get("suppliers")
    safe_suppliers = {}
    if isinstance(raw_suppliers, dict):
        for k, v in raw_suppliers.items():
            nk = normalize_sku(k)
            if not nk:
                continue
            safe_suppliers[nk] = str(v or "").strip()
    safe = {
        "updated_at": str((data or {}).get("updated_at") or ""),
        "items": _unique_skus((data or {}).get("items") or []),
        "suppliers": safe_suppliers,
        "warning": (data or {}).get("warning") if isinstance((data or {}).get("warning"), dict) else {"active": False},
        "snapshot_at": str((data or {}).get("snapshot_at") or ""),
    }
    with open(SKU_INDEX_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(safe, f, ensure_ascii=False, indent=2)

def load_sku_manual():
    try:
        if not os.path.exists(SKU_MANUAL_FILE):
            data = default_sku_manual()
            with open(SKU_MANUAL_FILE, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)
            return data
        with open(SKU_MANUAL_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, dict) and isinstance(data.get("items"), list):
                data["items"] = _unique_skus(data.get("items"))
                return data
    except Exception:
        pass
    return default_sku_manual()

def save_sku_manual(data):
    safe = {"items": _unique_skus((data or {}).get("items") or [])}
    with open(SKU_MANUAL_FILE, "w", encoding="utf-8") as f:
        json.dump(safe, f, ensure_ascii=False, indent=2)

def fetch_sku_index_from_sharepoint(*, use_existing_snapshot=False, guard_anomaly=True):
    import openpyxl

    snapshot_info = _snapshot_info_from_latest() if use_existing_snapshot else download_excel_snapshot()
    wb = openpyxl.load_workbook(snapshot_info["latest_path"], read_only=True, data_only=True)
    try:
        ws = None
        for name in wb.sheetnames:
            n = str(name).strip().lower().replace(" ", "").replace("-", "").replace("_", "")
            if "indeks" in n and "ean" in n:
                ws = wb[name]
                break
        if ws is None:
            raise RuntimeError("Brak arkusza 'Indeksy eany' w pliku SharePoint")

        producer_col_idx = None
        sku_col_idx = None
        header_row_idx = None
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            vals = [str(v).strip().lower() if v is not None else "" for v in row]
            normalized_vals = [re.sub(r"\s+", " ", v) for v in vals]
            if "producent" in normalized_vals:
                producer_col_idx = normalized_vals.index("producent")
            # Docelowo bierzemy SKU z kolumny "SYMBOL/ Index"
            if "symbol/ index" in normalized_vals:
                sku_col_idx = normalized_vals.index("symbol/ index")
                header_row_idx = i
                break
            if "symbol/index" in normalized_vals:
                sku_col_idx = normalized_vals.index("symbol/index")
                header_row_idx = i
                break
            if i >= 20:
                break

        if sku_col_idx is None:
            raise RuntimeError("Brak kolumny 'SYMBOL/ Index' w arkuszu 'Indeksy eany'")

        items = []
        supplier_votes = {}
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            if header_row_idx and i <= header_row_idx:
                continue
            if sku_col_idx >= len(row):
                continue
            sku = normalize_sku(row[sku_col_idx])
            if sku:
                items.append(sku)
                supplier = ""
                if producer_col_idx is not None and producer_col_idx < len(row):
                    supplier = str(row[producer_col_idx] or "").strip()
                if supplier:
                    if sku not in supplier_votes:
                        supplier_votes[sku] = {}
                    supplier_votes[sku][supplier] = supplier_votes[sku].get(supplier, 0) + 1

        suppliers = {}
        for sku, votes in supplier_votes.items():
            if not votes:
                continue
            suppliers[sku] = max(votes.items(), key=lambda x: x[1])[0]

        cache = {
            "updated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "items": _unique_skus(items),
            "suppliers": suppliers,
            "warning": {"active": False},
            "snapshot_at": snapshot_info.get("snapshot_at", ""),
        }
        if guard_anomaly:
            meta = _load_snapshot_meta()
            warning = _build_drop_warning(
                meta.get("last_good_sku_items", 0),
                len(cache["items"]),
                min_count=SKU_MIN_ITEMS,
                source_name="SKU",
            )
            if warning:
                _mark_snapshot_warning(warning)
                existing = load_sku_index_cache()
                existing["warning"] = warning
                return existing

        _mark_snapshot_ok(
            sku_items=len(cache["items"]),
            snapshot_file=snapshot_info.get("history_name") or _load_snapshot_meta().get("last_snapshot_file", ""),
        )
        save_sku_index_cache(cache)
        return cache
    finally:
        wb.close()

def get_merged_sku_index():
    excel_cache = load_sku_index_cache()
    if not (excel_cache.get("items") or []):
        try:
            excel_cache = fetch_sku_index_from_sharepoint()
        except Exception:
            pass
    manual = load_sku_manual()
    merged = _unique_skus((excel_cache.get("items") or []) + (manual.get("items") or []))
    suppliers = {}
    raw_suppliers = excel_cache.get("suppliers")
    if isinstance(raw_suppliers, dict):
        for k, v in raw_suppliers.items():
            nk = normalize_sku(k)
            if nk:
                suppliers[nk] = str(v or "").strip()
    return {
        "items": merged,
        "count": len(merged),
        "excel_count": len(_unique_skus(excel_cache.get("items") or [])),
        "manual_count": len(_unique_skus(manual.get("items") or [])),
        "updated_at": excel_cache.get("updated_at") or "",
        "suppliers": suppliers,
        "warning": excel_cache.get("warning") or _get_snapshot_warning(),
        "snapshot_at": excel_cache.get("snapshot_at") or "",
    }


def normalize_list_field(value, *, max_items=30, max_len=30):
    """
    Accepts:
      - list[str]
      - JSON string list (e.g. '["DG1","MH56"]')
      - plain string
    Returns (items: list[str], text: str)
    """
    items = []
    if value is None:
        return [], ""

    if isinstance(value, list):
        raw_items = value
    else:
        text = str(value).strip()
        if text.startswith("[") and text.endswith("]"):
            try:
                parsed = json.loads(text)
                raw_items = parsed if isinstance(parsed, list) else []
            except Exception:
                raw_items = []
        else:
            # Plain string; treat as single "text" field
            cleaned = text
            return [], cleaned

    for it in raw_items[:max_items]:
        if not isinstance(it, str):
            continue
        cleaned = it.strip()
        if not cleaned:
            continue
        if len(cleaned) > max_len:
            cleaned = cleaned[:max_len]
        if any(existing.lower() == cleaned.lower() for existing in items):
            continue
        items.append(cleaned)

    return items, ", ".join(items)

@app.context_processor
def inject_cache_buster():
    # Used in templates to bust long-lived static caching
    return {"cache_buster": CACHE_BUSTER}

# ========== USER MANAGEMENT FUNCTIONS ==========

def init_users_db():
    """Inicjalizuje tabelę users w bazie danych"""
    with closing(get_reklamacje_db()) as conn:
        if using_postgres():
            conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id TEXT PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'user',
                    is_active INTEGER NOT NULL DEFAULT 1,
                    must_change_password INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """)
        else:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS users (
                    id TEXT PRIMARY KEY,
                    username TEXT UNIQUE NOT NULL,
                    password_hash TEXT NOT NULL,
                    role TEXT NOT NULL DEFAULT 'user',
                    is_active INTEGER NOT NULL DEFAULT 1,
                    must_change_password INTEGER NOT NULL DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL
                )
            """)
        conn.commit()
        
        # Sprawdź czy istnieje konto admin, jeśli nie - utwórz
        admin_check = conn.execute(
            "SELECT id FROM users WHERE username = %s" if using_postgres() else "SELECT id FROM users WHERE username = ?",
            ('admin',)
        ).fetchone()
        
        if not admin_check:
            admin_id = uuid.uuid4().hex[:10]
            admin_password_hash = generate_password_hash('Admin123!')
            now_ts = datetime.datetime.utcnow().isoformat() + "Z"
            conn.execute(
                "INSERT INTO users (id, username, password_hash, role, is_active, must_change_password, created_at, updated_at) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
                if using_postgres()
                else "INSERT INTO users (id, username, password_hash, role, is_active, must_change_password, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
                (admin_id, 'admin', admin_password_hash, 'admin', 1, 1, now_ts, now_ts)
            )
            conn.commit()

def init_event_log_db():
    """Inicjalizuje tabelę dziennika zdarzeń"""
    with closing(get_reklamacje_db()) as conn:
        if using_postgres():
            conn.execute("""
                CREATE TABLE IF NOT EXISTS event_log (
                    id TEXT PRIMARY KEY,
                    created_at TEXT NOT NULL,
                    user_id TEXT,
                    username TEXT,
                    action TEXT NOT NULL,
                    details TEXT,
                    ip TEXT
                )
            """)
        else:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS event_log (
                    id TEXT PRIMARY KEY,
                    created_at TEXT NOT NULL,
                    user_id TEXT,
                    username TEXT,
                    action TEXT NOT NULL,
                    details TEXT,
                    ip TEXT
                )
            """)
        conn.commit()


def init_converter_db():
    """Inicjalizuje tabele dla modułu Konwerter (przechowywanie importów CSV)"""
    with closing(get_reklamacje_db()) as conn:
        if using_postgres():
            conn.execute("""
                CREATE TABLE IF NOT EXISTS converter_sessions (
                    id TEXT PRIMARY KEY,
                    user_id TEXT,
                    name TEXT NOT NULL,
                    source TEXT,
                    original_filename TEXT,
                    size_bytes INTEGER,
                    created_at TEXT NOT NULL,
                    csv_text TEXT NOT NULL
                )
            """)
        else:
            conn.execute("""
                CREATE TABLE IF NOT EXISTS converter_sessions (
                    id TEXT PRIMARY KEY,
                    user_id TEXT,
                    name TEXT NOT NULL,
                    source TEXT,
                    original_filename TEXT,
                    size_bytes INTEGER,
                    created_at TEXT NOT NULL,
                    csv_text TEXT NOT NULL
                )
            """)
        conn.commit()

def log_event(action, details=""):
    """Zapisuje zdarzenie do dziennika"""
    try:
        user_id = session.get('user_id')
        username = session.get('username')
        ip_addr = request.headers.get('X-Forwarded-For', request.remote_addr)
        event_id = uuid.uuid4().hex[:10]
        now_ts = datetime.datetime.utcnow().isoformat() + "Z"
        insert_sql = (
            "INSERT INTO event_log (id, created_at, user_id, username, action, details, ip) VALUES (%s, %s, %s, %s, %s, %s, %s)"
            if using_postgres()
            else "INSERT INTO event_log (id, created_at, user_id, username, action, details, ip) VALUES (?, ?, ?, ?, ?, ?, ?)"
        )
        with closing(get_reklamacje_db()) as conn:
            conn.execute(insert_sql, (event_id, now_ts, user_id, username, action, details, ip_addr))
            conn.commit()
    except Exception:
        pass

def hash_password(password):
    """Hashuje hasło używając werkzeug"""
    return generate_password_hash(password)

def validate_password_strength(password):
    """Sprawdza siłę hasła: min 9 znaków, duża litera, cyfra, znak specjalny"""
    if not password or len(password) < 9:
        return False
    has_upper = any(ch.isupper() for ch in password)
    has_digit = any(ch.isdigit() for ch in password)
    has_special = any(not ch.isalnum() for ch in password)
    return has_upper and has_digit and has_special

def verify_password(password_hash, password):
    """Weryfikuje hasło przeciwko hashowi"""
    return check_password_hash(password_hash, password)

def get_user(username):
    """Pobiera użytkownika z bazy danych"""
    with closing(get_reklamacje_db()) as conn:
        row = conn.execute(
            "SELECT id, username, password_hash, role, is_active, must_change_password, created_at, updated_at FROM users WHERE username = %s"
            if using_postgres()
            else "SELECT id, username, password_hash, role, is_active, must_change_password, created_at, updated_at FROM users WHERE username = ?",
            (username,)
        ).fetchone()
        if not row:
            return None
        return {
            "id": row["id"] if isinstance(row, dict) else row[0],
            "username": row["username"] if isinstance(row, dict) else row[1],
            "password_hash": row["password_hash"] if isinstance(row, dict) else row[2],
            "role": row["role"] if isinstance(row, dict) else row[3],
            "is_active": bool(row["is_active"] if isinstance(row, dict) else row[4]),
            "must_change_password": bool(row["must_change_password"] if isinstance(row, dict) else row[5]),
            "created_at": row["created_at"] if isinstance(row, dict) else row[6],
            "updated_at": row["updated_at"] if isinstance(row, dict) else row[7],
        }

def get_user_by_id(user_id):
    """Pobiera użytkownika po ID"""
    with closing(get_reklamacje_db()) as conn:
        row = conn.execute(
            "SELECT id, username, password_hash, role, is_active, must_change_password, created_at, updated_at FROM users WHERE id = %s"
            if using_postgres()
            else "SELECT id, username, password_hash, role, is_active, must_change_password, created_at, updated_at FROM users WHERE id = ?",
            (user_id,)
        ).fetchone()
        if not row:
            return None
        return {
            "id": row["id"] if isinstance(row, dict) else row[0],
            "username": row["username"] if isinstance(row, dict) else row[1],
            "password_hash": row["password_hash"] if isinstance(row, dict) else row[2],
            "role": row["role"] if isinstance(row, dict) else row[3],
            "is_active": bool(row["is_active"] if isinstance(row, dict) else row[4]),
            "must_change_password": bool(row["must_change_password"] if isinstance(row, dict) else row[5]),
            "created_at": row["created_at"] if isinstance(row, dict) else row[6],
            "updated_at": row["updated_at"] if isinstance(row, dict) else row[7],
        }

def create_user(username, password, role='user'):
    """Tworzy nowego użytkownika"""
    if not validate_password_strength(password):
        return None
    with closing(get_reklamacje_db()) as conn:
        # Sprawdź czy użytkownik już istnieje
        existing = conn.execute(
            "SELECT id FROM users WHERE username = %s" if using_postgres() else "SELECT id FROM users WHERE username = ?",
            (username,)
        ).fetchone()
        if existing:
            return None
        
        user_id = uuid.uuid4().hex[:10]
        password_hash = hash_password(password)
        now_ts = datetime.datetime.utcnow().isoformat() + "Z"
        conn.execute(
            "INSERT INTO users (id, username, password_hash, role, is_active, must_change_password, created_at, updated_at) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
            if using_postgres()
            else "INSERT INTO users (id, username, password_hash, role, is_active, must_change_password, created_at, updated_at) VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (user_id, username, password_hash, role, 1, 0, now_ts, now_ts)
        )
        conn.commit()
        return user_id

def update_user_password(user_id, new_password):
    """Aktualizuje hasło użytkownika"""
    if not validate_password_strength(new_password):
        return False
    with closing(get_reklamacje_db()) as conn:
        password_hash = hash_password(new_password)
        now_ts = datetime.datetime.utcnow().isoformat() + "Z"
        conn.execute(
            "UPDATE users SET password_hash = %s, must_change_password = %s, updated_at = %s WHERE id = %s"
            if using_postgres()
            else "UPDATE users SET password_hash = ?, must_change_password = ?, updated_at = ? WHERE id = ?",
            (password_hash, 0, now_ts, user_id)
        )
        conn.commit()
        return True

def update_user_username(user_id, new_username):
    """Aktualizuje nazwę użytkownika (unikalna)"""
    with closing(get_reklamacje_db()) as conn:
        existing = conn.execute(
            "SELECT id FROM users WHERE username = %s" if using_postgres() else "SELECT id FROM users WHERE username = ?",
            (new_username,)
        ).fetchone()
        if existing and (existing["id"] if isinstance(existing, dict) else existing[0]) != user_id:
            return False
        now_ts = datetime.datetime.utcnow().isoformat() + "Z"
        conn.execute(
            "UPDATE users SET username = %s, updated_at = %s WHERE id = %s"
            if using_postgres()
            else "UPDATE users SET username = ?, updated_at = ? WHERE id = ?",
            (new_username, now_ts, user_id)
        )
        conn.commit()
        return True

def update_user_role(user_id, role):
    """Aktualizuje rolę użytkownika (tylko admin, walidacja roli)"""
    if role not in ('admin', 'operator', 'user'):
        return False
    with closing(get_reklamacje_db()) as conn:
        now_ts = datetime.datetime.utcnow().isoformat() + "Z"
        conn.execute(
            "UPDATE users SET role = %s, updated_at = %s WHERE id = %s"
            if using_postgres()
            else "UPDATE users SET role = ?, updated_at = ? WHERE id = ?",
            (role, now_ts, user_id)
        )
        conn.commit()
        return True

def get_all_users():
    """Pobiera wszystkich użytkowników"""
    with closing(get_reklamacje_db()) as conn:
        rows = conn.execute(
            "SELECT id, username, role, is_active, must_change_password, created_at, updated_at FROM users ORDER BY username ASC"
        ).fetchall()
        return [
            {
                "id": row["id"] if isinstance(row, dict) else row[0],
                "username": row["username"] if isinstance(row, dict) else row[1],
                "role": row["role"] if isinstance(row, dict) else row[2],
                "is_active": bool(row["is_active"] if isinstance(row, dict) else row[3]),
                "must_change_password": bool(row["must_change_password"] if isinstance(row, dict) else row[4]),
                "created_at": row["created_at"] if isinstance(row, dict) else row[5],
                "updated_at": row["updated_at"] if isinstance(row, dict) else row[6],
            }
            for row in rows
        ]

def toggle_user_active(user_id):
    """Przełącza status aktywności użytkownika"""
    with closing(get_reklamacje_db()) as conn:
        # Pobierz aktualny status
        row = conn.execute(
            "SELECT is_active FROM users WHERE id = %s" if using_postgres() else "SELECT is_active FROM users WHERE id = ?",
            (user_id,)
        ).fetchone()
        if not row:
            return False
        current_status = bool(row["is_active"] if isinstance(row, dict) else row[0])
        new_status = 0 if current_status else 1
        now_ts = datetime.datetime.utcnow().isoformat() + "Z"
        conn.execute(
            "UPDATE users SET is_active = %s, updated_at = %s WHERE id = %s"
            if using_postgres()
            else "UPDATE users SET is_active = ?, updated_at = ? WHERE id = ?",
            (new_status, now_ts, user_id)
        )
        conn.commit()
        return True

def login_required(f):
    """Wymaga zalogowanego użytkownika, przekierowuje na /login jeśli brak sesji."""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get('logged_in'):
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def role_required(required_role):
    """Decorator wymagający określonej roli"""
    def decorator(f):
        @wraps(f)
        @login_required
        def decorated_function(*args, **kwargs):
            user_role = session.get('role', 'user')
            if required_role == 'admin' and user_role != 'admin':
                flash('Brak uprawnień do wykonania tej operacji.', 'error')
                return redirect(url_for('ustawienia'))
            return f(*args, **kwargs)
        return decorated_function
    return decorator

def get_product_stats():
    """Pobiera statystyki produktów z magazynu (pergole, domki, sztachety, deski)"""
    try:
        rows = fetch_inventory()
        data = parse_inventory_rows(rows)
        items = data.get('items', [])
        
        stats = {
            'pergole': 0,
            'domki': 0,
            'sztachety': 0,
            'deski': 0
        }
        
        for item in items:
            name_lower = (item.get('name') or '').lower()
            if 'pergol' in name_lower:
                stats['pergole'] += 1
            elif 'domek' in name_lower or 'domk' in name_lower:
                stats['domki'] += 1
            elif 'sztachet' in name_lower:
                stats['sztachety'] += 1
            elif 'desk' in name_lower or 'deska' in name_lower:
                stats['deski'] += 1
        
        return stats
    except Exception:
        # Jeśli nie można pobrać danych, zwróć zera
        return {'pergole': 0, 'domki': 0, 'sztachety': 0, 'deski': 0}

def get_reklamacje_count():
    """Pobiera liczbę zgłoszeń z bazy danych"""
    try:
        with closing(get_reklamacje_db()) as conn:
            if using_postgres():
                result = conn.execute("SELECT COUNT(1) AS cnt FROM reklamacje").fetchone()
            else:
                result = conn.execute("SELECT COUNT(1) AS cnt FROM reklamacje").fetchone()
            return result['cnt'] if result else 0
    except Exception:
        return 0

def get_reklamacje_stats():
    """Pobiera statystyki zgłoszeń: wszystkie, w trakcie, zakończone."""
    result = {
        "all": 0,
        "in_progress": 0,
        "closed": 0,
    }
    try:
        with closing(get_reklamacje_db()) as conn:
            total_row = conn.execute("SELECT COUNT(1) AS cnt FROM reklamacje").fetchone()
            in_progress_row = conn.execute(
                "SELECT COUNT(1) AS cnt FROM reklamacje WHERE status = %s" if using_postgres()
                else "SELECT COUNT(1) AS cnt FROM reklamacje WHERE status = ?",
                ("W trakcie",),
            ).fetchone()
            closed_row = conn.execute(
                "SELECT COUNT(1) AS cnt FROM reklamacje WHERE lower(trim(coalesce(status, ''))) = %s" if using_postgres()
                else "SELECT COUNT(1) AS cnt FROM reklamacje WHERE lower(trim(coalesce(status, ''))) = ?",
                ("zamknięta",),
            ).fetchone()
            result["all"] = int((total_row["cnt"] if total_row else 0) or 0)
            result["in_progress"] = int((in_progress_row["cnt"] if in_progress_row else 0) or 0)
            result["closed"] = int((closed_row["cnt"] if closed_row else 0) or 0)
    except Exception:
        pass
    return result

def _delivery_extract_product_type(opis):
    lc = (opis or "").lower()
    if "sztachet" in lc:
        return "Sztachety"
    if "pergola" in lc:
        return "Pergole"
    if "rolet" in lc:
        return "Rolety"
    if "panel" in lc:
        return "Panele"
    if "legar" in lc:
        return "Legary"
    if "listwa" in lc or "maskuj" in lc:
        return "Listwy"
    if "wspornik" in lc:
        return "Wsporniki"
    if "screen" in lc:
        return "Screen"
    if "louver" in lc or "wall" in lc:
        return "Louver Wall"
    if "solid" in lc:
        return "Solid"
    if "domek" in lc:
        return "Domki"
    if "deska" in lc or "dubline" in lc or "skogline" in lc or "vintage" in lc:
        return "Deski"
    return "Inne"

def _delivery_extract_brand(opis):
    raw = (opis or "").strip()
    lc = raw.lower()
    if "nordvic" in lc:
        return "Nordvic"
    if "nordeck" in lc:
        return "Nordeck"
    if "mirador" in lc or "80 solid" in lc:
        return "Mirador"
    return (raw.split()[0] if raw else "")

def _delivery_most_frequent(values):
    counts = {}
    for v in values:
        if not v:
            continue
        counts[v] = counts.get(v, 0) + 1
    if not counts:
        return ""
    return max(counts.items(), key=lambda x: x[1])[0]

def _parse_delivery_date(value):
    txt = str(value or "").strip()
    if not txt:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%Y/%m/%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.datetime.strptime(txt, fmt).date()
        except Exception:
            pass
    try:
        # fallback for broader ISO-like strings
        return datetime.date.fromisoformat(txt[:10])
    except Exception:
        return None

def _delivery_group_label(rows):
    brands = [_delivery_extract_brand(r.get("opis")) for r in rows]
    types = [_delivery_extract_product_type(r.get("opis")) for r in rows]
    main_brand = _delivery_most_frequent(brands)
    main_type = _delivery_most_frequent(types)
    if main_brand == "Nordvic" and (not main_type or main_type == "Inne"):
        main_type = "Pergole"
    if main_brand == "Mirador" and (not main_type or main_type == "Inne"):
        main_type = "Pergole"
    if main_brand == "Nordeck" and (not main_type or main_type == "Inne"):
        main_type = "Sztachety / Deski"
    if main_brand and main_type:
        return f"{main_brand} - {main_type}"
    return main_brand or main_type or "Dostawa"

def get_dashboard_delivery_stats(limit=3):
    result = {"upcoming": [], "completed": []}
    try:
        data = _load_dostawy_cache()
        rows = data.get("rows") or []
        if not rows:
            return result

        upcoming_by_date = {}
        completed_by_date = {}
        today = datetime.date.today()

        for r in rows:
            status = str(r.get("status") or "").lower()
            delivered_flag = ("dostarczon" in status) or ("zrealiz" in status)
            delivered_date_raw = r.get("dostarczono")
            if delivered_date_raw and str(delivered_date_raw).strip():
                delivered_flag = True

            planned_raw = str(r.get("planowana_dostawa") or "").strip()
            delivered_raw = str(r.get("dostarczono") or "").strip()
            date_raw = delivered_raw if delivered_flag and delivered_raw else planned_raw
            date_obj = _parse_delivery_date(date_raw) if date_raw else None
            if not date_raw:
                continue

            if delivered_flag:
                completed_by_date.setdefault(date_raw, []).append(r)
            else:
                if date_obj and date_obj < today:
                    continue
                upcoming_by_date.setdefault(date_raw, []).append(r)

        def build_items(grouped, reverse=False):
            prepared = []
            for date_raw, grp in grouped.items():
                date_obj = _parse_delivery_date(date_raw) or datetime.date.max
                prepared.append((date_obj, date_raw, grp))
            prepared.sort(key=lambda x: x[0], reverse=reverse)
            out = []
            for _, date_raw, grp in prepared[:limit]:
                out.append({
                    "name": _delivery_group_label(grp),
                    "date": date_raw,
                })
            return out

        result["upcoming"] = build_items(upcoming_by_date, reverse=False)
        result["completed"] = build_items(completed_by_date, reverse=True)
    except Exception:
        pass
    return result

@app.route('/')
@login_required
def index():
    stats = get_product_stats()
    reklamacje_count = get_reklamacje_count()
    reklamacje_stats = get_reklamacje_stats()
    delivery_stats = get_dashboard_delivery_stats()
    return render_template('dashboard.html', 
                         username=session.get('username', 'Użytkownik'),
                         stats=stats,
                         reklamacje_count=reklamacje_count,
                         reklamacje_stats=reklamacje_stats,
                         delivery_stats=delivery_stats)

# File sharing routes (authentication required)
@app.route('/upload', methods=['POST'])
@login_required
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'Brak pliku'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Nie wybrano pliku'}), 400

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)

        return jsonify({
            'success': True,
            'filename': filename,
            'url': f'/download/shared/{filename}',
            'message': f'Plik {filename} został przesłany pomyślnie!'
        }), 200
    else:
        return jsonify({'error': 'Niedozwolony typ pliku'}), 400

@app.route('/download/shared/<filename>')
@login_required
def download_shared_file(filename):
    """Pobierz plik z katalogu shared"""
    try:
        return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)
    except FileNotFoundError:
        return jsonify({'error': 'Plik nie istnieje'}), 404

@app.route('/files')
@login_required
def get_files():
    """Pobierz listę dostępnych plików"""
    files = []
    if os.path.exists(app.config['UPLOAD_FOLDER']):
        for filename in os.listdir(app.config['UPLOAD_FOLDER']):
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            if os.path.isfile(filepath):
                stat = os.stat(filepath)
                files.append({
                    'name': filename,
                    'size': stat.st_size,
                    'modified': datetime.datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
                })

    return jsonify(files)

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        remember = bool(request.form.get('remember'))
        
        user = get_user(username)
        if user and user['is_active'] and verify_password(user['password_hash'], password):
            session['logged_in'] = True
            session['username'] = user['username']
            session['user_id'] = user['id']
            session['role'] = user['role']
            session['must_change_password'] = user['must_change_password']
            session.permanent = remember
            
            # Jeśli użytkownik musi zmienić hasło, przekieruj do ustawień
            if user['must_change_password']:
                flash('Musisz zmienić hasło przed kontynuowaniem.', 'warning')
                log_event("login", "Wymuszona zmiana hasła")
                return redirect(url_for('ustawienia'))
            log_event("login", "Zalogowano poprawnie")
            return redirect(url_for('index'))
        else:
            return render_template('login.html', error='Nieprawidłowa nazwa użytkownika lub hasło')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    log_event("logout", "Wylogowano")
    session.clear()
    return redirect(url_for('login'))

@app.route('/magazyn')
@login_required
def magazyn():
    return render_template('magazyn_new.html',
                         sheet_id=GOOGLE_SHEET_ID,
                         sheet_gid=GOOGLE_SHEET_GID,
                         username=session.get('username', 'Użytkownik'))


def fetch_inventory():
    """
    Pobiera dane magazynowe i zwraca listę słowników.

    Priorytet:
      1) n8n (N8N_INVENTORY_URL) – JSON z danymi produktów
      2) Google Sheets / Drive (CSV) – dotychczasowe źródło (fallback)
    """
    last_error = None

    # 1) n8n webhook (preferowane źródło)
    if N8N_INVENTORY_URL:
        try:
            headers = {}
            if N8N_INVENTORY_API_KEY:
                headers["x-api-key"] = N8N_INVENTORY_API_KEY
            resp = requests.post(N8N_INVENTORY_URL, headers=headers, timeout=20)
            resp.raise_for_status()
            data = resp.json()

            # Obsłuż różne możliwe „kształty” odpowiedzi
            rows = None
            if isinstance(data, dict):
                rows = data.get("data") or data.get("rows")
            elif isinstance(data, list):
                # np. [{ "data": [...] }]
                if data and isinstance(data[0], dict) and "data" in data[0]:
                    rows = data[0]["data"]
                else:
                    rows = data

            if not isinstance(rows, list):
                raise ValueError("Nieoczekiwany format odpowiedzi n8n (brak listy danych)")
            if not rows:
                raise ValueError("Brak wierszy w odpowiedzi n8n")

            # Zakładamy, że pola w rows mają klucze zgodne z tymi używanymi
            # w parse_inventory_rows (Kod, Nazwa, jm, EAN, Stan, Rezerwacja, Dostępność).
            # Jeśli struktura jest inna, można ją tu zmapować.
            return rows
        except Exception as exc:
            last_error = exc

    # 2) Fallback: Google Sheets / Drive (CSV) – dotychczasowe zachowanie
    candidates = [
        f"https://docs.google.com/spreadsheets/d/{GOOGLE_SHEET_ID}/export?format=csv&gid={GOOGLE_SHEET_GID}",
        f"https://drive.google.com/uc?export=download&id={GOOGLE_SHEET_ID}",
    ]
    for url in candidates:
        try:
            resp = requests.get(url, timeout=15)
            resp.raise_for_status()
            content = resp.content.decode('utf-8', errors='ignore')
            # Odrzuć HTML (np. brak dostępu)
            if content.lstrip().lower().startswith("<!doctype") or content.lstrip().lower().startswith("<html"):
                raise ValueError("Otrzymano HTML zamiast CSV (brak dostępu?)")
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
            if rows:
                return rows
            last_error = ValueError("Brak wierszy w CSV")
        except Exception as exc:
            last_error = exc
            continue

    if last_error:
        raise last_error
    raise RuntimeError("Nie udało się pobrać danych z arkusza ani z n8n")


def parse_inventory_rows(rows):
    items = []
    now_ts = datetime.datetime.utcnow().isoformat() + "Z"
    for row in rows:
        def pick(*keys):
            for k in keys:
                if not k:
                    continue
                if k in row and row.get(k) not in (None, ""):
                    return row.get(k)
            return ""

        def as_str(value):
            if value is None:
                return ""
            return str(value).strip()

        items.append({
            # Kod / indeks
            "code": as_str(pick("Kod", "SKU", "ID_produktu")),
            # Nazwa produktu
            "name": as_str(pick("Nazwa")),
            # Jednostka miary
            "unit": as_str(pick("jm", "JM_sprzedazy")),
            # EAN
            "ean": as_str(pick("EAN")),
            # Stan magazynowy (aktualny)
            "stock": as_str(pick("Stan", "STAN", "Stan_magazynowy")),
            # Rezerwacja
            "reservation": as_str(pick("Rezerwacja")),
            # Dostępność (np. wolne sztuki)
            "availability": as_str(pick("Dostepnosc", "Dostępność")),
        })
    return {"updated_at": now_ts, "items": items}


def get_reklamacje_db():
    if REKLAMACJE_DB_URL:
        try:
            import psycopg
            from psycopg.rows import dict_row
        except Exception as exc:
            raise RuntimeError("Brak psycopg dla PostgreSQL") from exc
        conn = psycopg.connect(REKLAMACJE_DB_URL, row_factory=dict_row)
        return conn

    db_dir = os.path.dirname(REKLAMACJE_DB_PATH)
    if db_dir:
        os.makedirs(db_dir, exist_ok=True)
    conn = sqlite3.connect(REKLAMACJE_DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def using_postgres():
    return bool(REKLAMACJE_DB_URL)


def init_reklamacje_db():
    with closing(get_reklamacje_db()) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS reklamacje (
                id TEXT PRIMARY KEY,
                email TEXT,
                client TEXT,
                order_no TEXT NOT NULL,
                purchase_date TEXT,
                product_model TEXT NOT NULL,
                customer_name TEXT NOT NULL,
                phone TEXT NOT NULL,
                type TEXT NOT NULL,
                missing_or_damaged_desc TEXT NOT NULL,
                description TEXT,
                my_note TEXT,
                title TEXT,
                details TEXT,
                reported_at TEXT NOT NULL,
                status TEXT NOT NULL,
                decision TEXT,
                decision_reason TEXT,
                replacement_sent INTEGER NOT NULL DEFAULT 0,
                closed_at TEXT,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        if using_postgres():
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS email TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS client TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS order_no TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS purchase_date TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS product_model TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS customer_name TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS phone TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS type TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS missing_or_damaged_desc TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS description TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS my_note TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS title TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS details TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS reported_at TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS status TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS decision TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS decision_reason TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS replacement_sent INTEGER")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS closed_at TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS created_at TEXT")
            conn.execute("ALTER TABLE reklamacje ADD COLUMN IF NOT EXISTS updated_at TEXT")
            conn.execute("ALTER TABLE reklamacje ALTER COLUMN client DROP NOT NULL")
            conn.execute("ALTER TABLE reklamacje ALTER COLUMN title DROP NOT NULL")
        else:
            cols = [row["name"] if isinstance(row, dict) else row[1]
                    for row in conn.execute("PRAGMA table_info(reklamacje)").fetchall()]
            def ensure_col(name, ddl):
                if name not in cols:
                    conn.execute(ddl)
            ensure_col("email", "ALTER TABLE reklamacje ADD COLUMN email TEXT")
            ensure_col("client", "ALTER TABLE reklamacje ADD COLUMN client TEXT")
            ensure_col("order_no", "ALTER TABLE reklamacje ADD COLUMN order_no TEXT")
            ensure_col("purchase_date", "ALTER TABLE reklamacje ADD COLUMN purchase_date TEXT")
            ensure_col("product_model", "ALTER TABLE reklamacje ADD COLUMN product_model TEXT")
            ensure_col("customer_name", "ALTER TABLE reklamacje ADD COLUMN customer_name TEXT")
            ensure_col("phone", "ALTER TABLE reklamacje ADD COLUMN phone TEXT")
            ensure_col("type", "ALTER TABLE reklamacje ADD COLUMN type TEXT")
            ensure_col("missing_or_damaged_desc", "ALTER TABLE reklamacje ADD COLUMN missing_or_damaged_desc TEXT")
            ensure_col("description", "ALTER TABLE reklamacje ADD COLUMN description TEXT")
            ensure_col("my_note", "ALTER TABLE reklamacje ADD COLUMN my_note TEXT")
            ensure_col("title", "ALTER TABLE reklamacje ADD COLUMN title TEXT")
            ensure_col("details", "ALTER TABLE reklamacje ADD COLUMN details TEXT")
            ensure_col("reported_at", "ALTER TABLE reklamacje ADD COLUMN reported_at TEXT")
            ensure_col("status", "ALTER TABLE reklamacje ADD COLUMN status TEXT")
            ensure_col("decision", "ALTER TABLE reklamacje ADD COLUMN decision TEXT")
            ensure_col("decision_reason", "ALTER TABLE reklamacje ADD COLUMN decision_reason TEXT")
            ensure_col("replacement_sent", "ALTER TABLE reklamacje ADD COLUMN replacement_sent INTEGER")
            ensure_col("closed_at", "ALTER TABLE reklamacje ADD COLUMN closed_at TEXT")
            ensure_col("created_at", "ALTER TABLE reklamacje ADD COLUMN created_at TEXT")
            ensure_col("updated_at", "ALTER TABLE reklamacje ADD COLUMN updated_at TEXT")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS reklamacje_parts (
                id TEXT PRIMARY KEY,
                claim_id TEXT NOT NULL,
                part_name TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS reklamacje_attachments (
                id TEXT PRIMARY KEY,
                claim_id TEXT NOT NULL,
                file_path TEXT NOT NULL,
                original_name TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS reklamacje_notes (
                id TEXT PRIMARY KEY,
                claim_id TEXT NOT NULL,
                note_text TEXT NOT NULL,
                note_date TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS reklamacje_order_products (
                id TEXT PRIMARY KEY,
                claim_id TEXT NOT NULL,
                platform_id TEXT,
                product_sku TEXT,
                product_name TEXT NOT NULL,
                product_ean TEXT,
                quantity INTEGER NOT NULL DEFAULT 0,
                is_selected INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.commit()


def migrate_reklamacje_from_json():
    if not os.path.exists(REKLAMACJE_JSON_FILE):
        return
    try:
        with open(REKLAMACJE_JSON_FILE, "r", encoding="utf-8") as handle:
            data = json.load(handle)
        items = data if isinstance(data, list) else data.get("items", [])
    except Exception:
        return
    if not items:
        return
    with closing(get_reklamacje_db()) as conn:
        count_row = conn.execute("SELECT COUNT(1) AS cnt FROM reklamacje").fetchone()
        count = count_row["cnt"] if isinstance(count_row, dict) else count_row[0]
        if count:
            return
        for item in items:
            insert_sql = (
                """
                INSERT INTO reklamacje (
                    id, order_no, product_model, customer_name, email, phone, type,
                    missing_or_damaged_desc, description, reported_at, status,
                    decision, decision_reason, replacement_sent, closed_at,
                    purchase_date, created_at, updated_at
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                """
                if using_postgres()
                else """
                INSERT INTO reklamacje (
                    id, order_no, product_model, customer_name, email, phone, type,
                    missing_or_damaged_desc, description, reported_at, status,
                    decision, decision_reason, replacement_sent, closed_at,
                    purchase_date, created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """
            )
            conn.execute(
                insert_sql,
                (
                    item.get("id") or uuid.uuid4().hex[:10],
                    (item.get("order_no") or item.get("order") or "").strip() or "Brak",
                    (item.get("product_model") or "").strip() or "Pergola",
                    (item.get("customer_name") or item.get("client") or "").strip() or "Brak",
                    (item.get("email") or "").strip(),
                    (item.get("phone") or "").strip(),
                    (item.get("type") or REKLAMACJE_TYPES[0]).strip(),
                    (item.get("missing_or_damaged_desc") or item.get("details") or "").strip() or "Brak",
                    (item.get("description") or "").strip(),
                    item.get("reported_at") or datetime.datetime.utcnow().date().isoformat(),
                    (item.get("status") or REKLAMACJE_STATUSES[0]).strip(),
                    (item.get("decision") or "").strip(),
                    (item.get("decision_reason") or "").strip(),
                    1 if item.get("replacement_sent") else 0,
                    item.get("closed_at") or "",
                    item.get("purchase_date") or "",
                    item.get("created_at") or datetime.datetime.utcnow().isoformat() + "Z",
                    item.get("updated_at") or datetime.datetime.utcnow().isoformat() + "Z",
                ),
            )
        conn.commit()


def seed_reklamacje_if_empty():
    with closing(get_reklamacje_db()) as conn:
        count_row = conn.execute("SELECT COUNT(1) AS cnt FROM reklamacje").fetchone()
        count = count_row["cnt"] if isinstance(count_row, dict) else count_row[0]
        if count:
            return
        now = datetime.datetime.utcnow()
        demo = [
            {
                "id": uuid.uuid4().hex[:10],
                "order_no": "ZAM/1001",
                "purchase_date": (now - datetime.timedelta(days=14)).date().isoformat(),
                "product_model": "Pergola X (antracyt)",
                "customer_name": "Jan Kowalski",
                "email": "jan@example.com",
                "phone": "600123123",
                "type": "Uszkodzony element",
                "missing_or_damaged_desc": "Porysowana belka boczna",
                "description": "Uszkodzenie zauważone po rozpakowaniu.",
                "reported_at": (now - datetime.timedelta(days=5)).date().isoformat(),
                "status": "Nowa",
                "decision": "",
                "decision_reason": "",
                "replacement_sent": 0,
                "closed_at": "",
            },
            {
                "id": uuid.uuid4().hex[:10],
                "order_no": "ZAM/1002",
                "purchase_date": (now - datetime.timedelta(days=30)).date().isoformat(),
                "product_model": "Pergola Y (biała)",
                "customer_name": "Anna Nowak",
                "email": "anna@example.com",
                "phone": "501222333",
                "type": "Brak elementu",
                "missing_or_damaged_desc": "Brak śrub montażowych",
                "description": "Brakuje kompletu śrub do montażu.",
                "reported_at": (now - datetime.timedelta(days=12)).date().isoformat(),
                "status": "W trakcie",
                "decision": "Dosłanie części",
                "decision_reason": "Uzupełnienie brakujących elementów.",
                "replacement_sent": 1,
                "closed_at": "",
            },
        ]
        insert_sql = (
            """
            INSERT INTO reklamacje (
                id, order_no, purchase_date, product_model, customer_name, email, phone,
                type, missing_or_damaged_desc, description, reported_at, status, decision,
                decision_reason, replacement_sent, closed_at, created_at, updated_at
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            if using_postgres()
            else """
            INSERT INTO reklamacje (
                id, order_no, purchase_date, product_model, customer_name, email, phone,
                type, missing_or_damaged_desc, description, reported_at, status, decision,
                decision_reason, replacement_sent, closed_at, created_at, updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
        )
        now_ts = now.isoformat() + "Z"
        for item in demo:
            conn.execute(
                insert_sql,
                (
                    item["id"],
                    item["order_no"],
                    item["purchase_date"],
                    item["product_model"],
                    item["customer_name"],
                    item["email"],
                    item["phone"],
                    item["type"],
                    item["missing_or_damaged_desc"],
                    item["description"],
                    item["reported_at"],
                    item["status"],
                    item["decision"],
                    item["decision_reason"],
                    item["replacement_sent"],
                    item["closed_at"],
                    now_ts,
                    now_ts,
                ),
            )
        conn.commit()


try:
    init_reklamacje_db()
    init_users_db()
    init_event_log_db()
    init_converter_db()
    migrate_reklamacje_from_json()
    if os.environ.get("REKLAMACJE_SEED", "1") == "1":
        seed_reklamacje_if_empty()
except Exception as exc:
    print(f"[WARN] DB init failed: {exc}")


def reklamację_required_errors(payload):
    required = [
        ("order_no", "Numer zamówienia"),
        ("product_model", "Model / wariant produktu"),
        ("customer_name", "Imię i nazwisko"),
        ("email", "Email"),
        ("phone", "Telefon"),
        ("type", "Typ zgłoszenia"),
        ("missing_or_damaged_desc", "Co brakuje / co uszkodzone"),
        ("reported_at", "Data zgłoszenia"),
    ]
    errors = []
    for key, label in required:
        value = (payload.get(key) or "").strip()
        if not value:
            errors.append(label)
    return errors


def normalize_bool(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "tak", "yes", "y"}


def parse_date(value):
    if not value:
        return ""
    return str(value).strip()


def normalize_spaces(s):
    return re.sub(r"\s+", " ", str(s or "")).strip()


def order_product_key(p):
    # Klucz musi odpowiadać temu co generuje frontend w `reklamacje.html`.
    platform_id = normalize_spaces(p.get("platformId") or p.get("platform_id") or "")
    sku = normalize_spaces(p.get("sku") or p.get("product_sku") or p.get("productSku") or "")
    name = normalize_spaces(p.get("name") or p.get("product_name") or "")
    if platform_id:
        return f"id:{platform_id}"
    if sku:
        return f"sku:{sku}"
    if name:
        return f"name:{normalize_spaces(name).lower()}"
    return ""


def parse_order_products_payload(raw):
    """
    Frontend wysyła `order_products_payload` jako string JSON.
    Zwracamy: (products_list, selected_keys_set)
    """
    if raw is None:
        return [], set()
    payload = raw
    if isinstance(raw, str):
        if not raw.strip():
            return [], set()
        try:
            payload = json.loads(raw)
        except Exception:
            return [], set()
    if not isinstance(payload, dict):
        return [], set()

    raw_products = payload.get("products")
    if not isinstance(raw_products, list):
        raw_products = []

    raw_selected = payload.get("selectedKeys") or payload.get("selected_keys") or []
    if not isinstance(raw_selected, list):
        raw_selected = []
    selected_keys = {str(k) for k in raw_selected if str(k).strip()}

    products = []
    for p in raw_products[:80]:
        if not isinstance(p, dict):
            continue
        platform_id = normalize_spaces(p.get("platformId") or p.get("platform_id") or "")
        sku = normalize_spaces(p.get("sku") or p.get("product_sku") or p.get("productSku") or "")
        name = normalize_spaces(p.get("name") or p.get("product_name") or "")
        ean = normalize_spaces(p.get("ean") or p.get("product_ean") or p.get("productEan") or "")
        qty_raw = p.get("qty") if "qty" in p else p.get("quantity")
        qty = 0
        try:
            if qty_raw is not None and qty_raw != "":
                qty = int(float(str(qty_raw).replace(",", ".")))
        except Exception:
            qty = 0
        if not name and not sku:
            continue
        products.append({
            "platformId": platform_id,
            "sku": sku,
            "name": name,
            "ean": ean,
            "qty": qty,
        })

    return products, selected_keys


def compute_days_open(reported_at, closed_at):
    if not reported_at:
        return 0
    try:
        start = datetime.datetime.fromisoformat(reported_at[:10])
    except Exception:
        return 0
    if closed_at:
        try:
            end = datetime.datetime.fromisoformat(closed_at[:10])
        except Exception:
            end = datetime.datetime.utcnow()
    else:
        end = datetime.datetime.utcnow()
    return max(0, (end.date() - start.date()).days)


def fetch_parts(conn, claim_id):
    # Sprawdź czy kolumna sent istnieje
    try:
        conn.execute("SELECT sent FROM reklamacje_parts LIMIT 1")
        has_sent = True
    except:
        if using_postgres():
            try:
                conn.rollback()
            except Exception:
                pass
        has_sent = False
    select_sql = (
        "SELECT id, part_name, created_at" + (", sent" if has_sent else "") + " FROM reklamacje_parts WHERE claim_id = %s ORDER BY created_at ASC"
        if using_postgres()
        else "SELECT id, part_name, created_at" + (", sent" if has_sent else "") + " FROM reklamacje_parts WHERE claim_id = ? ORDER BY created_at ASC"
    )
    rows = conn.execute(select_sql, (claim_id,)).fetchall()
    return [
        {
            "id": row["id"] if isinstance(row, dict) else row[0],
            "part_name": row["part_name"] if isinstance(row, dict) else row[1],
            "created_at": row["created_at"] if isinstance(row, dict) else row[2],
            "sent": bool(row["sent"] if isinstance(row, dict) else (row[3] if len(row) > 3 else 0)) if has_sent else False,
        }
        for row in rows
    ]


def fetch_attachments(conn, claim_id):
    rows = conn.execute(
        """
        SELECT id, file_path, original_name, created_at
        FROM reklamacje_attachments
        WHERE claim_id = %s
        ORDER BY created_at ASC
        """ if using_postgres() else """
        SELECT id, file_path, original_name, created_at
        FROM reklamacje_attachments
        WHERE claim_id = ?
        ORDER BY created_at ASC
        """,
        (claim_id,),
    ).fetchall()
    return [
        {
            "id": row["id"],
            "file_path": row["file_path"],
            "original_name": row["original_name"],
            "created_at": row["created_at"],
        }
        for row in rows
    ]


def fetch_notes(conn, claim_id):
    rows = conn.execute(
        """
        SELECT id, note_text, note_date, created_at
        FROM reklamacje_notes
        WHERE claim_id = %s
        ORDER BY note_date DESC, created_at DESC
        """ if using_postgres() else """
        SELECT id, note_text, note_date, created_at
        FROM reklamacje_notes
        WHERE claim_id = ?
        ORDER BY note_date DESC, created_at DESC
        """,
        (claim_id,),
    ).fetchall()
    return [
        {
            "id": row["id"],
            "note_text": row["note_text"],
            "note_date": row["note_date"],
            "created_at": row["created_at"],
        }
        for row in rows
    ]


def fetch_order_products(conn, claim_id):
    rows = conn.execute(
        """
        SELECT platform_id, product_sku, product_name, product_ean, quantity, is_selected
        FROM reklamacje_order_products
        WHERE claim_id = %s
        ORDER BY created_at ASC
        """ if using_postgres() else """
        SELECT platform_id, product_sku, product_name, product_ean, quantity, is_selected
        FROM reklamacje_order_products
        WHERE claim_id = ?
        ORDER BY created_at ASC
        """,
        (claim_id,),
    ).fetchall()
    out = []
    for row in rows:
        platform_id = row["platform_id"]
        sku = row["product_sku"]
        name = row["product_name"]
        ean = row["product_ean"]
        qty = row["quantity"]
        selected = row["is_selected"]
        out.append({
            "platformId": platform_id or "",
            "sku": sku or "",
            "name": name or "",
            "ean": ean or "",
            "qty": int(qty) if qty is not None else 0,
            "selected": bool(selected),
        })
    return out


def save_order_products_for_claim(conn, claim_id, products, selected_keys):
    # Zawsze czyścimy dla claim_id i wstawiamy aktualny snapshot z formularza.
    delete_sql = (
        "DELETE FROM reklamacje_order_products WHERE claim_id = %s"
        if using_postgres()
        else "DELETE FROM reklamacje_order_products WHERE claim_id = ?"
    )

    insert_sql = (
        """
        INSERT INTO reklamacje_order_products (
            id, claim_id, platform_id, product_sku, product_name, product_ean,
            quantity, is_selected, created_at
        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
        """
        if using_postgres()
        else """
        INSERT INTO reklamacje_order_products (
            id, claim_id, platform_id, product_sku, product_name, product_ean,
            quantity, is_selected, created_at
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        """
    )

    now_ts = datetime.datetime.utcnow().isoformat() + "Z"
    selected_keys = selected_keys or set()

    conn.execute(delete_sql, (claim_id,))
    for p in products or []:
        if not isinstance(p, dict):
            continue
        name = normalize_spaces(p.get("name"))
        sku = normalize_spaces(p.get("sku"))
        platform_id = normalize_spaces(p.get("platformId"))
        ean = normalize_spaces(p.get("ean"))
        qty = p.get("qty", 0)
        try:
            qty_int = int(float(str(qty).replace(",", "."))) if qty is not None else 0
        except Exception:
            qty_int = 0

        if not name and not sku:
            continue

        key = order_product_key(p)
        is_selected = 1 if key in selected_keys else 0

        row_id = uuid.uuid4().hex[:10]
        conn.execute(
            insert_sql,
            (
                row_id,
                claim_id,
                platform_id,
                sku,
                name or sku,  # żeby spełnić NOT NULL (product_name)
                ean,
                qty_int,
                is_selected,
                now_ts,
            ),
        )



def fetch_last_notes(conn, claim_ids):
    if not claim_ids:
        return {}
    placeholders = ",".join(["%s"] * len(claim_ids)) if using_postgres() else ",".join(["?"] * len(claim_ids))
    rows = conn.execute(
        f"""
        SELECT claim_id, note_text, note_date, created_at
        FROM reklamacje_notes
        WHERE claim_id IN ({placeholders})
        ORDER BY note_date DESC, created_at DESC
        """,
        tuple(claim_ids),
    ).fetchall()
    latest = {}
    for row in rows:
        claim_id = row["claim_id"]
        if claim_id in latest:
            continue
        latest[claim_id] = {
            "note_text": row["note_text"],
            "note_date": row["note_date"],
        }
    return latest

def fetch_event_log(limit=200):
    with closing(get_reklamacje_db()) as conn:
        sql = (
            "SELECT created_at, user_id, username, action, details, ip FROM event_log ORDER BY created_at DESC LIMIT %s"
            if using_postgres()
            else "SELECT created_at, user_id, username, action, details, ip FROM event_log ORDER BY created_at DESC LIMIT ?"
        )
        rows = conn.execute(sql, (limit,)).fetchall()
    return [
        {
            "created_at": row["created_at"] if isinstance(row, dict) else row[0],
            "user_id": row["user_id"] if isinstance(row, dict) else row[1],
            "username": row["username"] if isinstance(row, dict) else row[2],
            "action": row["action"] if isinstance(row, dict) else row[3],
            "details": row["details"] if isinstance(row, dict) else row[4],
            "ip": row["ip"] if isinstance(row, dict) else row[5],
        }
        for row in rows
    ]

def build_event_log_filters(date_from=None, date_to=None, username=None, action=None, ip=None):
    where = []
    values = []
    if date_from:
        where.append("created_at >= %s" if using_postgres() else "created_at >= ?")
        values.append(f"{date_from}T00:00:00")
    if date_to:
        where.append("created_at <= %s" if using_postgres() else "created_at <= ?")
        values.append(f"{date_to}T23:59:59")
    if username:
        where.append("username = %s" if using_postgres() else "username = ?")
        values.append(username)
    if action:
        where.append("action = %s" if using_postgres() else "action = ?")
        values.append(action)
    if ip:
        where.append("ip = %s" if using_postgres() else "ip = ?")
        values.append(ip)
    return where, values

def fetch_event_log_filtered(date_from=None, date_to=None, username=None, action=None, ip=None, limit=2000):
    where, values = build_event_log_filters(date_from, date_to, username, action, ip)
    where_sql = f"WHERE {' AND '.join(where)}" if where else ""
    limit_sql = "LIMIT %s" if using_postgres() else "LIMIT ?"
    sql = f"SELECT created_at, user_id, username, action, details, ip FROM event_log {where_sql} ORDER BY created_at DESC {limit_sql}"
    values.append(limit)
    with closing(get_reklamacje_db()) as conn:
        rows = conn.execute(sql, tuple(values)).fetchall()
    return [
        {
            "created_at": row["created_at"] if isinstance(row, dict) else row[0],
            "user_id": row["user_id"] if isinstance(row, dict) else row[1],
            "username": row["username"] if isinstance(row, dict) else row[2],
            "action": row["action"] if isinstance(row, dict) else row[3],
            "details": row["details"] if isinstance(row, dict) else row[4],
            "ip": row["ip"] if isinstance(row, dict) else row[5],
        }
        for row in rows
    ]


@app.route('/api/reklamacje/<claim_id>/notes', methods=['POST'])
@login_required
def reklamacje_notes_api(claim_id):
    payload = request.get_json(silent=True) or {}
    note_text = (payload.get("note_text") or "").strip()
    if not note_text:
        return jsonify({"success": False, "error": "Wpisz treść uwagi."}), 400

    note_date = parse_date(payload.get("note_date")) or datetime.datetime.utcnow().date().isoformat()
    note_id = uuid.uuid4().hex[:10]
    now_ts = datetime.datetime.utcnow().isoformat() + "Z"

    insert_sql = (
        "INSERT INTO reklamacje_notes (id, claim_id, note_text, note_date, created_at) VALUES (%s, %s, %s, %s, %s)"
        if using_postgres()
        else "INSERT INTO reklamacje_notes (id, claim_id, note_text, note_date, created_at) VALUES (?, ?, ?, ?, ?)"
    )
    with closing(get_reklamacje_db()) as conn:
        conn.execute(insert_sql, (note_id, claim_id, note_text, note_date, now_ts))
        conn.commit()
        notes = fetch_notes(conn, claim_id)
    log_event("note_create", f"claim_id={claim_id} note_id={note_id}")
    return jsonify({"success": True, "notes": notes})


@app.route('/api/reklamacje/<claim_id>/notes/<note_id>', methods=['PUT', 'DELETE'])
@login_required
def reklamacje_notes_update_api(claim_id, note_id):
    if request.method == 'DELETE':
        delete_sql = (
            "DELETE FROM reklamacje_notes WHERE id = %s AND claim_id = %s"
            if using_postgres()
            else "DELETE FROM reklamacje_notes WHERE id = ? AND claim_id = ?"
        )
        with closing(get_reklamacje_db()) as conn:
            conn.execute(delete_sql, (note_id, claim_id))
            conn.commit()
            notes = fetch_notes(conn, claim_id)
        log_event("note_delete", f"claim_id={claim_id} note_id={note_id}")
        return jsonify({"success": True, "notes": notes})

    payload = request.get_json(silent=True) or {}
    note_text = (payload.get("note_text") or "").strip()
    if not note_text:
        return jsonify({"success": False, "error": "Wpisz treść uwagi."}), 400
    note_date = parse_date(payload.get("note_date")) or datetime.datetime.utcnow().date().isoformat()

    update_sql = (
        "UPDATE reklamacje_notes SET note_text = %s, note_date = %s WHERE id = %s AND claim_id = %s"
        if using_postgres()
        else "UPDATE reklamacje_notes SET note_text = ?, note_date = ? WHERE id = ? AND claim_id = ?"
    )
    with closing(get_reklamacje_db()) as conn:
        conn.execute(update_sql, (note_text, note_date, note_id, claim_id))
        conn.commit()
        notes = fetch_notes(conn, claim_id)
    log_event("note_update", f"claim_id={claim_id} note_id={note_id}")
    return jsonify({"success": True, "notes": notes})


@app.route('/api/reklamacje', methods=['GET', 'POST'])
@login_required
def reklamacje_api():
    if request.method == 'POST':
        payload = request.get_json(silent=True) or {}
        errors = reklamację_required_errors(payload)
        if errors:
            return jsonify({"success": False, "error": "Wymagane: " + ", ".join(errors)}), 400

        status = (payload.get("status") or REKLAMACJE_STATUSES[0]).strip()
        if status not in REKLAMACJE_STATUSES:
            status = REKLAMACJE_STATUSES[0]

        claim_id = uuid.uuid4().hex[:10]
        now_ts = datetime.datetime.utcnow().isoformat() + "Z"
        reported_at = parse_date(payload.get("reported_at")) or datetime.datetime.utcnow().date().isoformat()
        closed_at = parse_date(payload.get("closed_at"))
        if status == "Zamknięta" and not closed_at:
            closed_at = datetime.datetime.utcnow().date().isoformat()

        parts_items, _parts_text = normalize_list_field(payload.get("parts_list"))
        missing_items, missing_text = normalize_list_field(payload.get("missing_or_damaged_desc"))
        # Jeśli UI wysyła JSON-listę w polu "missing_or_damaged_desc", zapisujemy czytelny tekst.
        missing_value = missing_text if missing_text else (payload.get("missing_or_damaged_desc") or "")

        order_products_payload_provided = "order_products_payload" in payload
        order_products_products, order_products_selected_keys = ([], set())
        if order_products_payload_provided:
            order_products_products, order_products_selected_keys = parse_order_products_payload(payload.get("order_products_payload"))

        data = {
            "id": claim_id,
            "order_no": (payload.get("order_no") or "").strip(),
            "purchase_date": parse_date(payload.get("purchase_date")),
            "product_model": (payload.get("product_model") or "").strip(),
            "customer_name": (payload.get("customer_name") or "").strip(),
            "email": (payload.get("email") or "").strip(),
            "phone": (payload.get("phone") or "").strip(),
            "type": (payload.get("type") or "").strip(),
            "missing_or_damaged_desc": str(missing_value).strip(),
            "description": (payload.get("description") or "").strip(),
            "my_note": (payload.get("my_note") or "").strip(),
            "reported_at": reported_at,
            "status": status,
            "decision": (payload.get("decision") or "").strip(),
            "decision_reason": (payload.get("decision_reason") or "").strip(),
            "replacement_sent": 1 if normalize_bool(payload.get("replacement_sent")) else 0,
            "closed_at": closed_at or "",
            "created_at": now_ts,
            "updated_at": now_ts,
        }
        if not validate_email(data["email"]):
            return jsonify({"success": False, "error": "Nieprawidłowy adres email"}), 400
        if not validate_phone(data["phone"]):
            return jsonify({"success": False, "error": "Telefon może zawierać tylko cyfry"}), 400
        insert_sql = (
            """
            INSERT INTO reklamacje (
                id, order_no, purchase_date, product_model, customer_name, email, phone,
                type, missing_or_damaged_desc, description, my_note, reported_at, status, decision,
                decision_reason, replacement_sent, closed_at, created_at, updated_at,
                client, title, details
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            if using_postgres()
            else """
            INSERT INTO reklamacje (
                id, order_no, purchase_date, product_model, customer_name, email, phone,
                type, missing_or_damaged_desc, description, my_note, reported_at, status, decision,
                decision_reason, replacement_sent, closed_at, created_at, updated_at,
                client, title, details
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """
        )
        # Jeśli UI zbiera "części" na etapie tworzenia, zapisz je jako parts.
        # Fallback: jeśli parts_list puste, ale missing_or_damaged_desc było listą, użyj tego.
        parts_to_insert = parts_items or missing_items
        parts_insert_sql = (
            "INSERT INTO reklamacje_parts (id, claim_id, part_name, created_at) VALUES (%s, %s, %s, %s)"
            if using_postgres()
            else "INSERT INTO reklamacje_parts (id, claim_id, part_name, created_at) VALUES (?, ?, ?, ?)"
        )
        with closing(get_reklamacje_db()) as conn:
            conn.execute(
                insert_sql,
                (
                    data["id"],
                    data["order_no"],
                    data["purchase_date"],
                    data["product_model"],
                    data["customer_name"],
                    data["email"],
                    data["phone"],
                    data["type"],
                    data["missing_or_damaged_desc"],
                    data["description"],
                    data["my_note"],
                    data["reported_at"],
                    data["status"],
                    data["decision"],
                    data["decision_reason"],
                    data["replacement_sent"],
                    data["closed_at"],
                    data["created_at"],
                    data["updated_at"],
                    data["customer_name"],
                    data["type"],
                    data["description"],
                ),
            )
            if parts_to_insert:
                for part in parts_to_insert:
                    part_id = uuid.uuid4().hex[:10]
                    conn.execute(parts_insert_sql, (part_id, data["id"], part, now_ts))

            if order_products_payload_provided:
                save_order_products_for_claim(conn, data["id"], order_products_products, order_products_selected_keys)
            conn.commit()
        log_event("claim_create", f"claim_id={data['id']} order_no={data['order_no']}")
        data["days_open"] = compute_days_open(data["reported_at"], data["closed_at"])
        return jsonify({"success": True, "item": data}), 201

    status_filter = (request.args.get("status") or "").strip()
    type_filter = (request.args.get("type") or "").strip()
    query = (request.args.get("q") or "").strip().lower()
    sort = (request.args.get("sort") or "").strip()

    with closing(get_reklamacje_db()) as conn:
        rows = conn.execute(
            """
            SELECT
                id, order_no, purchase_date, product_model, customer_name, email, phone,
                type, missing_or_damaged_desc, description, my_note, reported_at, status, decision,
                decision_reason, replacement_sent, closed_at, created_at, updated_at
            FROM reklamacje
            ORDER BY created_at DESC
            """
        ).fetchall()

    claim_ids = []
    items = []
    for row in rows:
        item = {
            "id": row["id"],
            "order_no": row["order_no"] or "",
            "purchase_date": row["purchase_date"] or "",
            "product_model": row["product_model"] or "",
            "customer_name": row["customer_name"] or "",
            "email": row["email"] or "",
            "phone": row["phone"] or "",
            "type": row["type"] or "",
            "missing_or_damaged_desc": row["missing_or_damaged_desc"] or "",
            "description": row["description"] or "",
            "my_note": row["my_note"] or "",
            "reported_at": row["reported_at"] or "",
            "status": row["status"] or "",
            "decision": row["decision"] or "",
            "decision_reason": row["decision_reason"] or "",
            "replacement_sent": bool(row["replacement_sent"]),
            "closed_at": row["closed_at"] or "",
            "created_at": row["created_at"] or "",
            "updated_at": row["updated_at"] or "",
        }
        item["days_open"] = compute_days_open(item["reported_at"], item["closed_at"])
        items.append(item)
        claim_ids.append(item["id"])

    with closing(get_reklamacje_db()) as conn:
        last_notes = fetch_last_notes(conn, claim_ids)
    for item in items:
        note = last_notes.get(item["id"])
        item["last_note_text"] = note["note_text"] if note else ""
        item["last_note_date"] = note["note_date"] if note else ""

    if status_filter:
        items = [it for it in items if it["status"] == status_filter]
    if type_filter:
        items = [it for it in items if it["type"] == type_filter]
    if query:
        def match(it):
            hay = " ".join([
                it["order_no"], it["email"], it["customer_name"], it["product_model"], it["missing_or_damaged_desc"]
                , it["my_note"]
            ]).lower()
            return query in hay
        items = [it for it in items if match(it)]

    if sort == "days_open_desc":
        items.sort(key=lambda it: it["days_open"], reverse=True)
    else:
        # Domyślnie: najpierw otwarte (najnowsze na górze), potem zamknięte (też od najnowszych)
        open_items = []
        closed_items = []
        for it in items:
            s = (it["status"] or "").strip().lower()
            if s == "zamknięta":
                closed_items.append(it)
            else:
                open_items.append(it)
        items = open_items + closed_items

    return jsonify({
        "success": True,
        "updated_at": datetime.datetime.utcnow().isoformat() + "Z",
        "items": items,
        "statuses": REKLAMACJE_STATUSES,
        "types": REKLAMACJE_TYPES,
        "decisions": REKLAMACJE_DECISIONS,
    })


@app.route('/api/reklamacje/<claim_id>', methods=['GET', 'PUT', 'DELETE'])
@login_required
def reklamacje_detail_api(claim_id):
    if request.method == 'DELETE':
        # Usuń zgłoszenie i powiązane dane
        with closing(get_reklamacje_db()) as conn:
            # Usuń załączniki (pliki)
            attachments_sql = (
                "SELECT file_path FROM reklamacje_attachments WHERE claim_id = %s"
                if using_postgres()
                else "SELECT file_path FROM reklamacje_attachments WHERE claim_id = ?"
            )
            attachments = conn.execute(attachments_sql, (claim_id,)).fetchall()
            for att in attachments:
                file_path = att["file_path"] if isinstance(att, dict) else att[0]
                try:
                    if os.path.exists(file_path):
                        os.remove(file_path)
                except Exception:
                    pass
            
            # Usuń z bazy danych
            delete_attachments_sql = (
                "DELETE FROM reklamacje_attachments WHERE claim_id = %s"
                if using_postgres()
                else "DELETE FROM reklamacje_attachments WHERE claim_id = ?"
            )
            delete_parts_sql = (
                "DELETE FROM reklamacje_parts WHERE claim_id = %s"
                if using_postgres()
                else "DELETE FROM reklamacje_parts WHERE claim_id = ?"
            )
            delete_order_products_sql = (
                "DELETE FROM reklamacje_order_products WHERE claim_id = %s"
                if using_postgres()
                else "DELETE FROM reklamacje_order_products WHERE claim_id = ?"
            )
            delete_claim_sql = (
                "DELETE FROM reklamacje WHERE id = %s"
                if using_postgres()
                else "DELETE FROM reklamacje WHERE id = ?"
            )
            conn.execute(delete_attachments_sql, (claim_id,))
            conn.execute(delete_parts_sql, (claim_id,))
            conn.execute(delete_order_products_sql, (claim_id,))
            conn.execute(delete_claim_sql, (claim_id,))
            conn.commit()
        
        # Usuń katalog z plikami
        upload_dir = Path(app.config["REKLAMACJE_UPLOADS"]) / claim_id
        try:
            if upload_dir.exists():
                import shutil
                shutil.rmtree(upload_dir)
        except Exception:
            pass
        log_event("claim_delete", f"claim_id={claim_id}")
        return jsonify({"success": True})
    
    if request.method == 'PUT':
        payload = request.get_json(silent=True) or {}
        status = (payload.get("status") or "").strip()
        if status and status not in REKLAMACJE_STATUSES:
            return jsonify({"success": False, "error": "Nieprawidłowy status"}), 400

        email = (payload.get("email") or "").strip()
        phone = (payload.get("phone") or "").strip()
        if email and not validate_email(email):
            return jsonify({"success": False, "error": "Nieprawidłowy adres email"}), 400
        if phone and not validate_phone(phone):
            return jsonify({"success": False, "error": "Telefon może zawierać tylko cyfry"}), 400

        reported_at = parse_date(payload.get("reported_at"))
        closed_at = parse_date(payload.get("closed_at"))

        order_products_payload_provided = "order_products_payload" in payload
        order_products_products, order_products_selected_keys = ([], set())
        if order_products_payload_provided:
            order_products_products, order_products_selected_keys = parse_order_products_payload(payload.get("order_products_payload"))

        now_ts = datetime.datetime.utcnow().isoformat() + "Z"
        select_status_sql = (
            "SELECT status FROM reklamacje WHERE id = %s"
            if using_postgres()
            else "SELECT status FROM reklamacje WHERE id = ?"
        )
        update_sql = (
            """
            UPDATE reklamacje SET
                order_no = %s,
                purchase_date = %s,
                product_model = %s,
                customer_name = %s,
                email = %s,
                phone = %s,
                type = %s,
                missing_or_damaged_desc = %s,
                description = %s,
                my_note = %s,
                reported_at = %s,
                status = %s,
                decision = %s,
                decision_reason = %s,
                replacement_sent = %s,
                closed_at = %s,
                updated_at = %s,
                client = %s,
                title = %s,
                details = %s
            WHERE id = %s
            """
            if using_postgres()
            else """
            UPDATE reklamacje SET
                order_no = ?,
                purchase_date = ?,
                product_model = ?,
                customer_name = ?,
                email = ?,
                phone = ?,
                type = ?,
                missing_or_damaged_desc = ?,
                description = ?,
                my_note = ?,
                reported_at = ?,
                status = ?,
                decision = ?,
                decision_reason = ?,
                replacement_sent = ?,
                closed_at = ?,
                updated_at = ?,
                client = ?,
                title = ?,
                details = ?
            WHERE id = ?
            """
        )
        with closing(get_reklamacje_db()) as conn:
            existing = conn.execute(select_status_sql, (claim_id,)).fetchone()
            existing_status = ""
            if isinstance(existing, dict):
                existing_status = str(existing.get("status") or "").strip()
            elif existing:
                existing_status = str(existing[0] or "").strip()

            final_status = status or existing_status or REKLAMACJE_STATUSES[0]
            if final_status not in REKLAMACJE_STATUSES:
                final_status = REKLAMACJE_STATUSES[0]
            if final_status == "Zamknięta" and not closed_at:
                closed_at = datetime.datetime.utcnow().date().isoformat()

            conn.execute(
                update_sql,
                (
                    (payload.get("order_no") or "").strip(),
                    parse_date(payload.get("purchase_date")),
                    (payload.get("product_model") or "").strip(),
                    (payload.get("customer_name") or "").strip(),
                    email,
                    phone,
                    (payload.get("type") or "").strip(),
                    (payload.get("missing_or_damaged_desc") or "").strip(),
                    (payload.get("description") or "").strip(),
                    (payload.get("my_note") or "").strip(),
                    reported_at,
                    final_status,
                    (payload.get("decision") or "").strip(),
                    (payload.get("decision_reason") or "").strip(),
                    1 if normalize_bool(payload.get("replacement_sent")) else 0,
                    closed_at or "",
                    now_ts,
                    (payload.get("customer_name") or "").strip(),
                    (payload.get("type") or "").strip(),
                    (payload.get("description") or "").strip(),
                    claim_id,
                ),
            )

            if order_products_payload_provided:
                save_order_products_for_claim(conn, claim_id, order_products_products, order_products_selected_keys)
            conn.commit()
        log_event("claim_update", f"claim_id={claim_id} status={final_status}")
        return jsonify({"success": True})

    with closing(get_reklamacje_db()) as conn:
        row = conn.execute(
            """
            SELECT
                id, order_no, purchase_date, product_model, customer_name, email, phone,
                type, missing_or_damaged_desc, description, my_note, reported_at, status, decision,
                decision_reason, replacement_sent, closed_at, created_at, updated_at
            FROM reklamacje
            WHERE id = %s
            """ if using_postgres() else """
            SELECT
                id, order_no, purchase_date, product_model, customer_name, email, phone,
                type, missing_or_damaged_desc, description, my_note, reported_at, status, decision,
                decision_reason, replacement_sent, closed_at, created_at, updated_at
            FROM reklamacje
            WHERE id = ?
            """,
            (claim_id,),
        ).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Nie znaleziono zgłoszenia"}), 404
        item = {
            "id": row["id"],
            "order_no": row["order_no"] or "",
            "purchase_date": row["purchase_date"] or "",
            "product_model": row["product_model"] or "",
            "customer_name": row["customer_name"] or "",
            "email": row["email"] or "",
            "phone": row["phone"] or "",
            "type": row["type"] or "",
            "missing_or_damaged_desc": row["missing_or_damaged_desc"] or "",
            "description": row["description"] or "",
            "my_note": row["my_note"] or "",
            "reported_at": row["reported_at"] or "",
            "status": row["status"] or "",
            "decision": row["decision"] or "",
            "decision_reason": row["decision_reason"] or "",
            "replacement_sent": bool(row["replacement_sent"]),
            "closed_at": row["closed_at"] or "",
            "created_at": row["created_at"] or "",
            "updated_at": row["updated_at"] or "",
        }
        item["days_open"] = compute_days_open(item["reported_at"], item["closed_at"])
        item["parts"] = fetch_parts(conn, claim_id)
        item["attachments"] = fetch_attachments(conn, claim_id)
        item["notes"] = fetch_notes(conn, claim_id)
        item["order_products"] = fetch_order_products(conn, claim_id)
    return jsonify({"success": True, "item": item})


@app.route('/api/reklamacje/<claim_id>/parts', methods=['POST'])
@login_required
def reklamacje_parts_api(claim_id):
    payload = request.get_json(silent=True) or {}
    part_name = (payload.get("part_name") or "").strip()
    if not part_name:
        return jsonify({"success": False, "error": "Wymagana nazwa części"}), 400
    part_id = uuid.uuid4().hex[:10]
    now_ts = datetime.datetime.utcnow().isoformat() + "Z"
    sent = 1 if normalize_bool(payload.get("sent")) else 0
    with closing(get_reklamacje_db()) as conn:
        # Sprawdź czy kolumna sent istnieje, jeśli nie - dodaj
        try:
            conn.execute("SELECT sent FROM reklamacje_parts LIMIT 1")
        except:
            if using_postgres():
                try:
                    conn.rollback()
                except Exception:
                    pass
            try:
                conn.execute("ALTER TABLE reklamacje_parts ADD COLUMN sent INTEGER NOT NULL DEFAULT 0")
                conn.commit()
            except:
                pass
        insert_sql = (
            "INSERT INTO reklamacje_parts (id, claim_id, part_name, sent, created_at) VALUES (%s, %s, %s, %s, %s)"
            if using_postgres()
            else "INSERT INTO reklamacje_parts (id, claim_id, part_name, sent, created_at) VALUES (?, ?, ?, ?, ?)"
        )
        conn.execute(insert_sql, (part_id, claim_id, part_name, sent, now_ts))
        conn.commit()
    log_event("part_add", f"claim_id={claim_id} part_id={part_id} sent={bool(sent)}")
    return jsonify({"success": True, "item": {"id": part_id, "part_name": part_name, "sent": bool(sent), "created_at": now_ts}}), 201


@app.route('/api/reklamacje/<claim_id>/parts/<part_id>', methods=['PUT', 'DELETE'])
@login_required
def reklamacje_parts_update_api(claim_id, part_id):
    if request.method == 'PUT':
        payload = request.get_json(silent=True) or {}
        sent = 1 if normalize_bool(payload.get("sent")) else 0
        with closing(get_reklamacje_db()) as conn:
            # Sprawdź czy kolumna sent istnieje, jeśli nie - dodaj
            try:
                conn.execute("SELECT sent FROM reklamacje_parts LIMIT 1")
            except:
                if using_postgres():
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                try:
                    conn.execute("ALTER TABLE reklamacje_parts ADD COLUMN sent INTEGER NOT NULL DEFAULT 0")
                    conn.commit()
                except:
                    pass
            update_sql = (
                "UPDATE reklamacje_parts SET sent = %s WHERE id = %s AND claim_id = %s"
                if using_postgres()
                else "UPDATE reklamacje_parts SET sent = ? WHERE id = ? AND claim_id = ?"
            )
            conn.execute(update_sql, (sent, part_id, claim_id))
            conn.commit()
        log_event("part_update", f"claim_id={claim_id} part_id={part_id} sent={bool(sent)}")
        return jsonify({"success": True})
    else:  # DELETE
        delete_sql = (
            "DELETE FROM reklamacje_parts WHERE id = %s AND claim_id = %s"
            if using_postgres()
            else "DELETE FROM reklamacje_parts WHERE id = ? AND claim_id = ?"
        )
        with closing(get_reklamacje_db()) as conn:
            conn.execute(delete_sql, (part_id, claim_id))
            conn.commit()
        log_event("part_delete", f"claim_id={claim_id} part_id={part_id}")
        return jsonify({"success": True})


@app.route('/api/reklamacje/<claim_id>/attachments', methods=['GET', 'POST'])
@login_required
def reklamacje_attachments_api(claim_id):
    if request.method == 'GET':
        # Pobierz listę załączników dla zgłoszenia
        with closing(get_reklamacje_db()) as conn:
            attachments = fetch_attachments(conn, claim_id)
        return jsonify({"success": True, "items": attachments})
    
    # POST - dodaj załączniki
    files = request.files.getlist("files")
    if not files:
        return jsonify({"success": False, "error": "Brak plików"}), 400
    saved = []
    upload_root = Path(app.config["REKLAMACJE_UPLOADS"]) / claim_id
    upload_root.mkdir(parents=True, exist_ok=True)
    now_ts = datetime.datetime.utcnow().isoformat() + "Z"
    insert_sql = (
        "INSERT INTO reklamacje_attachments (id, claim_id, file_path, original_name, created_at) VALUES (%s, %s, %s, %s, %s)"
        if using_postgres()
        else "INSERT INTO reklamacje_attachments (id, claim_id, file_path, original_name, created_at) VALUES (?, ?, ?, ?, ?)"
    )
    with closing(get_reklamacje_db()) as conn:
        for file in files:
            if not file or not file.filename:
                continue
            filename = secure_filename(file.filename)
            if not allowed_file(filename):
                continue
            file_id = uuid.uuid4().hex[:10]
            stored_name = f"{file_id}_{filename}"
            file_path = str(upload_root / stored_name)
            file.save(file_path)
            # Kompresja obrazów po zapisie (best-effort).
            compress_uploaded_image_in_place(file_path)
            conn.execute(insert_sql, (file_id, claim_id, file_path, filename, now_ts))
            saved.append({"id": file_id, "original_name": filename, "created_at": now_ts})
        conn.commit()
    if saved:
        log_event("attachment_add", f"claim_id={claim_id} count={len(saved)}")
    return jsonify({"success": True, "items": saved})


@app.route('/api/reklamacje/<claim_id>/attachments/<attachment_id>', methods=['GET'])
@login_required
def reklamacje_attachment_download(claim_id, attachment_id):
    select_sql = (
        "SELECT file_path, original_name FROM reklamacje_attachments WHERE id = %s AND claim_id = %s"
        if using_postgres()
        else "SELECT file_path, original_name FROM reklamacje_attachments WHERE id = ? AND claim_id = ?"
    )
    with closing(get_reklamacje_db()) as conn:
        row = conn.execute(select_sql, (attachment_id, claim_id)).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Brak pliku"}), 404
        file_path = row["file_path"]
        original_name = row["original_name"]
    directory = os.path.dirname(file_path)
    filename = os.path.basename(file_path)
    return send_from_directory(directory, filename, as_attachment=True, download_name=original_name)


@app.route('/api/reklamacje/<claim_id>/attachments/<attachment_id>', methods=['DELETE'])
@login_required
def reklamacje_attachment_delete(claim_id, attachment_id):
    select_sql = (
        "SELECT file_path FROM reklamacje_attachments WHERE id = %s AND claim_id = %s"
        if using_postgres()
        else "SELECT file_path FROM reklamacje_attachments WHERE id = ? AND claim_id = ?"
    )
    delete_sql = (
        "DELETE FROM reklamacje_attachments WHERE id = %s AND claim_id = %s"
        if using_postgres()
        else "DELETE FROM reklamacje_attachments WHERE id = ? AND claim_id = ?"
    )
    with closing(get_reklamacje_db()) as conn:
        row = conn.execute(select_sql, (attachment_id, claim_id)).fetchone()
        if not row:
            return jsonify({"success": False, "error": "Brak pliku"}), 404
        conn.execute(delete_sql, (attachment_id, claim_id))
        conn.commit()
    try:
        os.remove(row["file_path"])
    except FileNotFoundError:
        pass
    log_event("attachment_delete", f"claim_id={claim_id} attachment_id={attachment_id}")
    return jsonify({"success": True})


@app.route('/api/magazyn')
@login_required
def magazyn_api():
    try:
        rows = fetch_inventory()
        data = parse_inventory_rows(rows)
        return jsonify({"success": True, **data})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route('/api/kartony', methods=['GET', 'PUT'])
@login_required
def kartony_api():
    if request.method == 'GET':
        return jsonify({"success": True, "data": load_kartony_data()})

    payload = request.get_json(silent=True) or {}
    data = payload.get("data")
    if not isinstance(data, dict) or not isinstance(data.get("products"), list):
        return jsonify({"success": False, "error": "Zły format. Oczekuję obiektu z polem products: []"}), 400
    try:
        save_kartony_data(data)
        products_count = len(data.get("products", [])) if isinstance(data, dict) else 0
        log_event("kartony_update", f"products={products_count}")
        return jsonify({"success": True})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500

@app.route('/api/planer', methods=['GET', 'PUT'])
@login_required
def planer_api():
    if request.method == 'GET':
        return jsonify({"success": True, "data": load_planer_data()})

    payload = request.get_json(silent=True) or {}
    data = payload.get("data")
    rows = data.get("rows") if isinstance(data, dict) else None
    if not isinstance(rows, list):
        return jsonify({"success": False, "error": "Zły format. Oczekuję obiektu z polem rows: []"}), 400

    normalized_rows = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        rid = str(row.get("id") or "").strip()[:40]
        if not rid:
            continue
        normalized_rows.append({
            "id": rid,
            "supplier": str(row.get("supplier") or "").strip()[:120],
            "sku": str(row.get("sku") or "").strip()[:120],
            "unit": str(row.get("unit") or "").strip()[:40],
            "project_qty": str(row.get("project_qty") or "").strip()[:40],
        })

    try:
        save_planer_data({"rows": normalized_rows})
        log_event("planer_update", f"rows={len(normalized_rows)}")
        return jsonify({"success": True})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500

@app.route('/api/sku-index', methods=['GET'])
@login_required
def api_sku_index():
    merged = get_merged_sku_index()
    return jsonify({
        "success": True,
        "items": merged["items"],
        "suppliers": merged.get("suppliers", {}),
        "count": merged["count"],
        "excel_count": merged["excel_count"],
        "manual_count": merged["manual_count"],
        "updated_at": merged["updated_at"],
        "snapshot_at": merged.get("snapshot_at", ""),
        "source": "cache+manual",
        "warning": merged.get("warning") or {"active": False},
    })

@app.route('/api/sku-index/refresh', methods=['POST'])
@login_required
def api_sku_index_refresh():
    try:
        cache = fetch_sku_index_from_sharepoint()
        merged = get_merged_sku_index()
        log_event("sku_index_refresh", f"excel_count={len(cache.get('items') or [])}")
        return jsonify({
            "success": True,
            "items": merged["items"],
            "suppliers": merged.get("suppliers", {}),
            "count": merged["count"],
            "excel_count": merged["excel_count"],
            "manual_count": merged["manual_count"],
            "updated_at": merged["updated_at"],
            "snapshot_at": merged.get("snapshot_at", ""),
            "source": "sharepoint+manual",
            "warning": merged.get("warning") or {"active": False},
        })
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500

@app.route('/api/sku-index/manual', methods=['POST'])
@login_required
def api_sku_index_manual_add():
    payload = request.get_json(silent=True) or {}
    incoming = []
    if isinstance(payload.get("sku"), str):
        incoming.append(payload.get("sku"))
    if isinstance(payload.get("items"), list):
        incoming.extend(payload.get("items"))

    new_items = _unique_skus(incoming)
    if not new_items:
        return jsonify({"success": False, "error": "Brak SKU do dodania"}), 400

    manual = load_sku_manual()
    current = _unique_skus(manual.get("items") or [])
    merged_manual = _unique_skus(current + new_items)
    save_sku_manual({"items": merged_manual})

    merged = get_merged_sku_index()
    log_event("sku_manual_add", f"added={len(new_items)} manual_total={len(merged_manual)}")
    return jsonify({
        "success": True,
        "added": new_items,
        "items": merged["items"],
        "suppliers": merged.get("suppliers", {}),
        "count": merged["count"],
        "excel_count": merged["excel_count"],
        "manual_count": merged["manual_count"],
        "updated_at": merged["updated_at"],
        "snapshot_at": merged.get("snapshot_at", ""),
        "source": "cache+manual",
        "warning": merged.get("warning") or {"active": False},
    })

@app.route('/api/sku-index/manual/<path:sku_value>', methods=['DELETE'])
@login_required
def api_sku_index_manual_delete(sku_value):
    target = normalize_sku(sku_value)
    if not target:
        return jsonify({"success": False, "error": "Brak SKU do usunięcia"}), 400

    manual = load_sku_manual()
    current = _unique_skus(manual.get("items") or [])
    if target not in current:
        return jsonify({"success": False, "error": "Nie znaleziono SKU na liście ręcznej"}), 404

    updated = [s for s in current if s != target]
    save_sku_manual({"items": updated})
    merged = get_merged_sku_index()
    log_event("sku_manual_delete", f"sku={target} manual_total={len(updated)}")
    return jsonify({
        "success": True,
        "removed": target,
        "items": merged["items"],
        "suppliers": merged.get("suppliers", {}),
        "count": merged["count"],
        "excel_count": merged["excel_count"],
        "manual_count": merged["manual_count"],
        "updated_at": merged["updated_at"],
        "snapshot_at": merged.get("snapshot_at", ""),
        "source": "cache+manual",
        "warning": merged.get("warning") or {"active": False},
    })

@app.route('/tickets')
@login_required
def tickets():
    return render_template('tickets.html', username=session.get('username', 'Użytkownik'))

@app.route('/availability')
@login_required
def availability():
    return render_template('availability.html', 
                         sheet_id=GOOGLE_SHEET_ID,
                         sheet_gid=GOOGLE_SHEET_GID,
                         username=session.get('username', 'Użytkownik'))

DOSTAWY_XLSX_URL = "https://suuhouse-my.sharepoint.com/personal/k_lubbe_suuhouse_pl/_layouts/15/download.aspx?share=IQBNN6em8fEtSJQtjukpEPX-AYmtyVkHqs8muDavntTA4ys"
DOSTAWY_CACHE_FILE = os.path.join(os.path.dirname(__file__), "dostawy_cache.json")
DOSTAWY_SNAPSHOT_DIR = os.environ.get("DOSTAWY_SNAPSHOT_DIR", os.path.join(os.path.dirname(__file__), "excel_snapshots"))
DOSTAWY_SNAPSHOT_HISTORY_DIR = os.path.join(DOSTAWY_SNAPSHOT_DIR, "history")
DOSTAWY_SNAPSHOT_META_FILE = os.path.join(DOSTAWY_SNAPSHOT_DIR, "snapshot_meta.json")
DOSTAWY_SNAPSHOT_LATEST_FILE = os.path.join(DOSTAWY_SNAPSHOT_DIR, "latest.xlsx")
DOSTAWY_SECONDARY_BACKUP_DIR = os.environ.get("DOSTAWY_SECONDARY_BACKUP_DIR", "/opt/batgraf/backups/mag1_excel")
DOSTAWY_HISTORY_RETENTION = int(os.environ.get("DOSTAWY_HISTORY_RETENTION", "60") or "60")
DOSTAWY_DROP_THRESHOLD = float(os.environ.get("DOSTAWY_DROP_THRESHOLD", "0.6") or "0.6")
DOSTAWY_MIN_ROWS = int(os.environ.get("DOSTAWY_MIN_ROWS", "50") or "50")
SKU_MIN_ITEMS = int(os.environ.get("SKU_MIN_ITEMS", "50") or "50")


def _default_snapshot_meta():
    return {
        "last_snapshot_at": "",
        "last_snapshot_file": "",
        "last_good_snapshot_file": "",
        "last_good_dostawy_rows": 0,
        "last_good_sku_items": 0,
        "warning": None,
    }


def _load_snapshot_meta():
    try:
        if os.path.exists(DOSTAWY_SNAPSHOT_META_FILE) and os.path.getsize(DOSTAWY_SNAPSHOT_META_FILE) > 2:
            with open(DOSTAWY_SNAPSHOT_META_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
                if isinstance(data, dict):
                    out = _default_snapshot_meta()
                    out.update(data)
                    return out
    except Exception:
        pass
    return _default_snapshot_meta()


def _save_snapshot_meta(meta):
    os.makedirs(DOSTAWY_SNAPSHOT_DIR, exist_ok=True)
    safe = _default_snapshot_meta()
    if isinstance(meta, dict):
        safe.update(meta)
    with open(DOSTAWY_SNAPSHOT_META_FILE, "w", encoding="utf-8") as f:
        json.dump(safe, f, ensure_ascii=False, indent=2)


def _ensure_snapshot_dirs():
    os.makedirs(DOSTAWY_SNAPSHOT_DIR, exist_ok=True)
    os.makedirs(DOSTAWY_SNAPSHOT_HISTORY_DIR, exist_ok=True)
    secondary_ready = True
    try:
        os.makedirs(DOSTAWY_SECONDARY_BACKUP_DIR, exist_ok=True)
    except Exception:
        secondary_ready = False
    return secondary_ready


def _sha256_of_file(path):
    h = hashlib.sha256()
    with open(path, "rb") as f:
        while True:
            chunk = f.read(65536)
            if not chunk:
                break
            h.update(chunk)
    return h.hexdigest()


def _rotate_snapshot_history():
    try:
        files = []
        for name in os.listdir(DOSTAWY_SNAPSHOT_HISTORY_DIR):
            if not name.lower().endswith(".xlsx"):
                continue
            fp = os.path.join(DOSTAWY_SNAPSHOT_HISTORY_DIR, name)
            try:
                mtime = os.path.getmtime(fp)
            except Exception:
                mtime = 0
            files.append((mtime, fp))
        files.sort(key=lambda x: x[0], reverse=True)
        for _, fp in files[DOSTAWY_HISTORY_RETENTION:]:
            try:
                os.remove(fp)
            except Exception:
                pass
    except Exception:
        pass


def _fix_snapshot_file_mode(path):
    try:
        if os.path.exists(path):
            os.chmod(path, 0o660)
    except Exception:
        pass


def _cleanup_tmp_snapshots():
    try:
        for name in os.listdir(DOSTAWY_SNAPSHOT_DIR):
            if not name.startswith("tmp_dostawy_") or not name.endswith(".xlsx"):
                continue
            fp = os.path.join(DOSTAWY_SNAPSHOT_DIR, name)
            try:
                os.remove(fp)
            except Exception:
                pass
    except Exception:
        pass


def _build_drop_warning(prev_count, new_count, *, min_count, source_name):
    prev = int(prev_count or 0)
    new = int(new_count or 0)
    if new < min_count:
        return {
            "source": source_name,
            "active": True,
            "type": "low_count",
            "message": f"Wykryto niską liczbę rekordów ({new}) dla {source_name}.",
            "previous_count": prev,
            "new_count": new,
            "detected_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        }
    if prev > 0:
        limit = int(prev * (1.0 - DOSTAWY_DROP_THRESHOLD))
        if new < max(1, limit):
            drop_pct = round((1.0 - (new / prev)) * 100.0, 1)
            return {
                "source": source_name,
                "active": True,
                "type": "sudden_drop",
                "message": f"Wykryto gwałtowny spadek danych ({prev} -> {new}, spadek {drop_pct}%).",
                "previous_count": prev,
                "new_count": new,
                "drop_percent": drop_pct,
                "detected_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            }
    return None


def _get_snapshot_warning():
    warning = (_load_snapshot_meta().get("warning") or {})
    if not isinstance(warning, dict):
        return {"active": False}
    if warning.get("active"):
        return warning
    return {"active": False}


def _mark_snapshot_warning(warning):
    meta = _load_snapshot_meta()
    meta["warning"] = warning if isinstance(warning, dict) else {"active": False}
    _save_snapshot_meta(meta)


def _mark_snapshot_ok(*, dostawy_rows=None, sku_items=None, snapshot_file=None):
    meta = _load_snapshot_meta()
    if isinstance(dostawy_rows, int) and dostawy_rows >= 0:
        meta["last_good_dostawy_rows"] = dostawy_rows
    if isinstance(sku_items, int) and sku_items >= 0:
        meta["last_good_sku_items"] = sku_items
    if snapshot_file:
        meta["last_good_snapshot_file"] = snapshot_file
    meta["warning"] = {"active": False}
    _save_snapshot_meta(meta)


def _download_sharepoint_bytes():
    last_err = None
    for attempt in range(3):
        try:
            resp = requests.get(DOSTAWY_XLSX_URL, timeout=30)
            resp.raise_for_status()
            return resp.content
        except Exception as exc:
            last_err = exc
            if attempt < 2:
                time.sleep(1.2 * (attempt + 1))
    raise last_err if last_err else RuntimeError("Nie udało się pobrać pliku SharePoint")


def download_excel_snapshot():
    """Pobiera i zapisuje snapshot Excela: latest + history + kopia secondary."""
    import openpyxl

    secondary_ready = _ensure_snapshot_dirs()
    content = _download_sharepoint_bytes()
    now = datetime.datetime.now()
    stamp = now.strftime("%Y%m%d_%H%M%S")
    history_name = f"dostawy_{stamp}.xlsx"
    tmp_path = os.path.join(DOSTAWY_SNAPSHOT_DIR, f"tmp_{history_name}")
    history_path = os.path.join(DOSTAWY_SNAPSHOT_HISTORY_DIR, history_name)

    with open(tmp_path, "wb") as f:
        f.write(content)

    wb = openpyxl.load_workbook(tmp_path, read_only=True, data_only=True)
    wb.close()

    shutil.copy2(tmp_path, history_path)
    _fix_snapshot_file_mode(history_path)
    os.replace(tmp_path, DOSTAWY_SNAPSHOT_LATEST_FILE)
    _fix_snapshot_file_mode(DOSTAWY_SNAPSHOT_LATEST_FILE)

    secondary_path = ""
    secondary_error = ""
    if secondary_ready:
        try:
            secondary_path = os.path.join(DOSTAWY_SECONDARY_BACKUP_DIR, history_name)
            shutil.copy2(history_path, secondary_path)
            _fix_snapshot_file_mode(secondary_path)
        except Exception as exc:
            secondary_error = str(exc)

    _cleanup_tmp_snapshots()

    file_sha = _sha256_of_file(DOSTAWY_SNAPSHOT_LATEST_FILE)
    size_bytes = os.path.getsize(DOSTAWY_SNAPSHOT_LATEST_FILE)
    meta = _load_snapshot_meta()
    meta["last_snapshot_at"] = now.strftime("%Y-%m-%d %H:%M:%S")
    meta["last_snapshot_file"] = history_name
    _save_snapshot_meta(meta)
    _rotate_snapshot_history()

    return {
        "snapshot_at": meta["last_snapshot_at"],
        "history_name": history_name,
        "latest_path": DOSTAWY_SNAPSHOT_LATEST_FILE,
        "latest_sha256": file_sha,
        "latest_size_bytes": size_bytes,
        "secondary_path": secondary_path,
        "secondary_error": secondary_error,
        "warning": _get_snapshot_warning(),
    }


def _snapshot_info_from_latest():
    if not os.path.exists(DOSTAWY_SNAPSHOT_LATEST_FILE):
        return download_excel_snapshot()
    meta = _load_snapshot_meta()
    return {
        "snapshot_at": meta.get("last_snapshot_at") or "",
        "history_name": meta.get("last_snapshot_file") or "",
        "latest_path": DOSTAWY_SNAPSHOT_LATEST_FILE,
        "latest_sha256": _sha256_of_file(DOSTAWY_SNAPSHOT_LATEST_FILE),
        "latest_size_bytes": os.path.getsize(DOSTAWY_SNAPSHOT_LATEST_FILE),
        "secondary_path": "",
        "secondary_error": "",
        "warning": _get_snapshot_warning(),
    }

def _fetch_dostawy_xlsx(*, use_existing_snapshot=False, guard_anomaly=True):
    """Download/parse Excel from snapshot and cache as JSON."""
    import openpyxl
    snapshot_info = _snapshot_info_from_latest() if use_existing_snapshot else download_excel_snapshot()
    wb = openpyxl.load_workbook(snapshot_info["latest_path"], read_only=True, data_only=True)
    # Główne dane dostaw trzymamy w arkuszu „DOSTAWY_2026”
    if "DOSTAWY_2026" in wb.sheetnames:
        ws = wb["DOSTAWY_2026"]
    else:
        ws = wb[wb.sheetnames[0]]
    rows_out = []
    headers = None
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = row
            continue
        vals = list(row)
        if not any(vals):
            continue
        def fmt(v):
            if v is None:
                return ""
            if isinstance(v, datetime.datetime):
                return v.strftime("%Y-%m-%d")
            return str(v)
        # Nowa struktura: 15 kolumn
        # 0: Dostawca
        # 1: nr transportu
        # 2: nr kontenera
        # 3: SYMBOL/ Index
        # 4: EAN
        # 5: OPIS2
        # 6: ilość szt.
        # 7: Ilość palet
        # 8: planowana dostawa
        # 9: nr proformy
        # 10: Nr faktury
        # 11: Dostarczono
        # 12: koszty net PLN
        # 13: koszt Brutto allin /szt
        # 14: status
        rows_out.append({
            "dostawca": fmt(vals[0]) if len(vals) > 0 else "",
            "nr_transportu": fmt(vals[1]) if len(vals) > 1 else "",
            "symbol": fmt(vals[3]) if len(vals) > 3 else "",
            "ean": fmt(vals[4]) if len(vals) > 4 else "",
            "opis": fmt(vals[5]) if len(vals) > 5 else "",
            # ilość sztuk – jako string (może być liczba lub pusty)
            "ilosc": fmt(vals[6]) if len(vals) > 6 else "",
            "ilosc_palet": fmt(vals[7]) if len(vals) > 7 else "",
            "planowana_dostawa": fmt(vals[8]) if len(vals) > 8 else "",
            "nr_proformy": fmt(vals[9]) if len(vals) > 9 else "",
            "nr_faktury": fmt(vals[10]) if len(vals) > 10 else "",
            "dostarczono": fmt(vals[11]) if len(vals) > 11 else "",
            "koszty_allin": fmt(vals[12]) if len(vals) > 12 else "",
            "koszt_netto_szt": fmt(vals[13]) if len(vals) > 13 else "",
            "status": fmt(vals[14]) if len(vals) > 14 else "",
        })
    news_out = []
    if "Aktualności" in wb.sheetnames:
        ws_news = wb["Aktualności"]
        for i, row in enumerate(ws_news.iter_rows(values_only=True)):
            if i == 0:
                continue
            vals = list(row)
            if not any(vals):
                continue
            date_val = vals[0] if len(vals) > 0 else None
            text_val = vals[1] if len(vals) > 1 else ""
            if not text_val:
                continue
            d = ""
            if isinstance(date_val, datetime.datetime):
                d = date_val.strftime("%Y-%m-%d %H:%M")
            elif date_val:
                d = str(date_val)
            news_out.append({"date": d, "text": str(text_val)})

    wb.close()
    cache = {
        "last_update": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "rows": rows_out,
        "news": news_out,
        "warning": {"active": False},
        "source_snapshot_at": snapshot_info.get("snapshot_at") or "",
        "source_snapshot_file": snapshot_info.get("history_name") or "",
    }
    if guard_anomaly:
        meta = _load_snapshot_meta()
        warning = _build_drop_warning(
            meta.get("last_good_dostawy_rows", 0),
            len(rows_out),
            min_count=DOSTAWY_MIN_ROWS,
            source_name="Dostawy",
        )
        if warning:
            _mark_snapshot_warning(warning)
            if os.path.exists(DOSTAWY_CACHE_FILE) and os.path.getsize(DOSTAWY_CACHE_FILE) > 2:
                with open(DOSTAWY_CACHE_FILE, "r", encoding="utf-8") as f:
                    old_cache = json.load(f)
                old_cache["warning"] = warning
                old_cache["source_snapshot_at"] = snapshot_info.get("snapshot_at") or old_cache.get("source_snapshot_at", "")
                old_cache["source_snapshot_file"] = snapshot_info.get("history_name") or old_cache.get("source_snapshot_file", "")
                return old_cache
            cache["warning"] = warning
        else:
            _mark_snapshot_ok(
                dostawy_rows=len(rows_out),
                snapshot_file=snapshot_info.get("history_name") or _load_snapshot_meta().get("last_snapshot_file", ""),
            )
    else:
        _mark_snapshot_ok(
            dostawy_rows=len(rows_out),
            snapshot_file=snapshot_info.get("history_name") or _load_snapshot_meta().get("last_snapshot_file", ""),
        )

    with open(DOSTAWY_CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False)
    return cache

def _load_dostawy_cache():
    if os.path.exists(DOSTAWY_CACHE_FILE) and os.path.getsize(DOSTAWY_CACHE_FILE) > 2:
        with open(DOSTAWY_CACHE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        data["warning"] = data.get("warning") or _get_snapshot_warning()
        return data
    return _fetch_dostawy_xlsx()

@app.route('/api/dostawy')
@login_required
def api_dostawy():
    try:
        data = _load_dostawy_cache()
        data["warning"] = data.get("warning") or _get_snapshot_warning()
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e), "rows": [], "last_update": None, "warning": _get_snapshot_warning()}), 500

@app.route('/api/dostawy/refresh', methods=['POST'])
@login_required
def api_dostawy_refresh():
    try:
        data = _fetch_dostawy_xlsx()
        data["warning"] = data.get("warning") or _get_snapshot_warning()
        log_event("dostawy_snapshot_ok", f"rows={len(data.get('rows') or [])} warning={bool(data.get('warning', {}).get('active'))}")
        return jsonify(data)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route('/api/dostawy/health')
@login_required
def api_dostawy_health():
    meta = _load_snapshot_meta()
    warning = meta.get("warning") if isinstance(meta.get("warning"), dict) else {"active": False}
    return jsonify({
        "success": True,
        "last_snapshot_at": meta.get("last_snapshot_at") or "",
        "last_snapshot_file": meta.get("last_snapshot_file") or "",
        "last_good_snapshot_file": meta.get("last_good_snapshot_file") or "",
        "last_good_dostawy_rows": int(meta.get("last_good_dostawy_rows") or 0),
        "last_good_sku_items": int(meta.get("last_good_sku_items") or 0),
        "warning": warning if warning else {"active": False},
    })


@app.route('/api/dostawy/source-file')
@login_required
def api_dostawy_source_file():
    try:
        if not os.path.exists(DOSTAWY_SNAPSHOT_LATEST_FILE):
            _snapshot_info_from_latest()
        if not os.path.exists(DOSTAWY_SNAPSHOT_LATEST_FILE):
            return jsonify({"success": False, "error": "Brak lokalnej kopii Excel"}), 404
        log_event("dostawy_download_source", "latest.xlsx")
        return send_from_directory(DOSTAWY_SNAPSHOT_DIR, "latest.xlsx", as_attachment=True, download_name="dostawy_latest.xlsx")
    except PermissionError:
        return jsonify({"success": False, "error": "Brak uprawnień zapisu do katalogu snapshotów na serwerze."}), 500
    except Exception as exc:
        return jsonify({"success": False, "error": f"Nie udało się przygotować kopii: {exc}"}), 500


@app.route('/api/dostawy/restore-last-good', methods=['POST'])
@login_required
def api_dostawy_restore_last_good():
    meta = _load_snapshot_meta()
    filename = str(meta.get("last_good_snapshot_file") or "").strip()
    if not filename:
        return jsonify({"success": False, "error": "Brak zapisanej działającej wersji"}), 404
    src = os.path.join(DOSTAWY_SNAPSHOT_HISTORY_DIR, filename)
    if not os.path.exists(src):
        return jsonify({"success": False, "error": "Nie znaleziono pliku ostatniej działającej wersji"}), 404
    os.makedirs(DOSTAWY_SNAPSHOT_DIR, exist_ok=True)
    shutil.copy2(src, DOSTAWY_SNAPSHOT_LATEST_FILE)
    data = _fetch_dostawy_xlsx(use_existing_snapshot=True, guard_anomaly=False)
    fetch_sku_index_from_sharepoint(use_existing_snapshot=True, guard_anomaly=False)
    _mark_snapshot_warning({"active": False})
    log_event("dostawy_restore_last_good", f"snapshot={filename}")
    return jsonify({
        "success": True,
        "message": "Przywrócono ostatnią działającą wersję.",
        "snapshot_file": filename,
        "data": data,
    })

@app.route('/dostawy')
@login_required
def dostawy():
    return render_template('dostawy.html', username=session.get('username', 'Użytkownik'))

@app.route('/rozmiary')
@login_required
def rozmiary():
    return render_template('rozmiary.html', username=session.get('username', 'Użytkownik'))

@app.route('/magazyn-new')
@login_required
def magazyn_new():
    return render_template('magazyn_new.html', username=session.get('username', 'Użytkownik'))

@app.route('/reklamacje')
@login_required
def reklamacje():
    return render_template('reklamacje.html', username=session.get('username', 'Użytkownik'))

@app.route('/kartony')
@login_required
def kartony():
    return render_template('kartony.html', username=session.get('username', 'Użytkownik'))

@app.route('/konwerter')
@login_required
def konwerter():
    return render_template('konwerter.html', username=session.get('username', 'Użytkownik'))


def create_converter_session(user_id, name, original_filename, csv_text, source=None):
    """Tworzy nową sesję konwertera i zapisuje surowy CSV w bazie"""
    session_id = uuid.uuid4().hex[:10]
    created_at = datetime.datetime.utcnow().isoformat() + "Z"
    size_bytes = len(csv_text.encode("utf-8"))
    with closing(get_reklamacje_db()) as conn:
        conn.execute(
            "INSERT INTO converter_sessions (id, user_id, name, source, original_filename, size_bytes, created_at, csv_text) "
            "VALUES (%s, %s, %s, %s, %s, %s, %s, %s)"
            if using_postgres()
            else "INSERT INTO converter_sessions (id, user_id, name, source, original_filename, size_bytes, created_at, csv_text) "
                 "VALUES (?, ?, ?, ?, ?, ?, ?, ?)",
            (session_id, user_id, name, source, original_filename, size_bytes, created_at, csv_text),
        )
        conn.commit()
    return session_id


def list_converter_sessions_for_user(user_id):
    """Zwraca listę sesji konwertera dla użytkownika"""
    with closing(get_reklamacje_db()) as conn:
        rows = conn.execute(
            "SELECT id, name, source, original_filename, size_bytes, created_at "
            "FROM converter_sessions WHERE user_id = %s ORDER BY created_at DESC"
            if using_postgres()
            else "SELECT id, name, source, original_filename, size_bytes, created_at "
                 "FROM converter_sessions WHERE user_id = ? ORDER BY created_at DESC",
            (user_id,),
        ).fetchall()
    result = []
    for row in rows:
        # row może być dict (Postgres) albo tuple (SQLite)
        if isinstance(row, dict):
            result.append(
                {
                    "id": row["id"],
                    "name": row["name"],
                    "source": row.get("source"),
                    "original_filename": row.get("original_filename"),
                    "size_bytes": row.get("size_bytes") or 0,
                    "created_at": row.get("created_at"),
                }
            )
        else:
            result.append(
                {
                    "id": row[0],
                    "name": row[1],
                    "source": row[2],
                    "original_filename": row[3],
                    "size_bytes": row[4] or 0,
                    "created_at": row[5],
                }
            )
    return result


def get_converter_session_csv(user_id, session_id):
    """Pobiera surowy CSV dla danej sesji użytkownika"""
    with closing(get_reklamacje_db()) as conn:
        row = conn.execute(
            "SELECT id, user_id, name, original_filename, csv_text FROM converter_sessions WHERE id = %s"
            if using_postgres()
            else "SELECT id, user_id, name, original_filename, csv_text FROM converter_sessions WHERE id = ?",
            (session_id,),
        ).fetchone()
    if not row:
        return None
    row_user_id = row["user_id"] if isinstance(row, dict) else row[1]
    if str(row_user_id) != str(user_id):
        return None
    if isinstance(row, dict):
        return {
            "id": row["id"],
            "user_id": row["user_id"],
            "name": row["name"],
            "original_filename": row.get("original_filename"),
            "csv_text": row["csv_text"],
        }
    return {
        "id": row[0],
        "user_id": row[1],
        "name": row[2],
        "original_filename": row[3],
        "csv_text": row[4],
    }


def delete_converter_session(user_id, session_id):
    """Usuwa sesję konwertera (tylko właściciel)"""
    with closing(get_reklamacje_db()) as conn:
        row = conn.execute(
            "SELECT user_id FROM converter_sessions WHERE id = %s"
            if using_postgres()
            else "SELECT user_id FROM converter_sessions WHERE id = ?",
            (session_id,),
        ).fetchone()
        if not row:
            return False
        row_user_id = row["user_id"] if isinstance(row, dict) else row[0]
        if str(row_user_id) != str(user_id):
            return False
        conn.execute(
            "DELETE FROM converter_sessions WHERE id = %s"
            if using_postgres()
            else "DELETE FROM converter_sessions WHERE id = ?",
            (session_id,),
        )
        conn.commit()
        return True

@app.route('/ustawienia')
@login_required
def ustawienia():
    return render_template('ustawienia.html', username=session.get('username', 'Użytkownik'))

@app.route('/planer')
@login_required
def planer():
    return render_template('planer.html', username=session.get('username', 'Użytkownik'))


@app.route('/kontakty')
@login_required
def kontakty():
    return render_template('kontakty.html', username=session.get('username', 'Użytkownik'))


def _kontakty_parse_request():
    ct = (request.content_type or "").lower()
    if "multipart/form-data" in ct:
        return {
            **_normalize_kontakt_fields(dict(request.form)),
            "logo": request.files.get("logo"),
            "remove_logo": (request.form.get("remove_logo") or "").strip().lower() in ("1", "true", "on", "yes"),
        }
    payload = request.get_json(silent=True) or {}
    return {
        **_normalize_kontakt_fields(payload),
        "logo": None,
        "remove_logo": bool(payload.get("remove_logo")),
    }


@app.route('/api/kontakty', methods=['GET', 'POST'])
@login_required
def api_kontakty():
    if request.method == 'GET':
        data = load_kontakty_data()
        return jsonify({"success": True, "contacts": data.get("contacts", [])})

    parsed = _kontakty_parse_request()
    fields = {k: parsed[k] for k in ("full_name", "company", "phone", "email")}
    if not any(fields.values()):
        return jsonify({"success": False, "error": "Uzupełnij co najmniej jedno pole (np. imię i nazwisko lub firmę)."}), 400

    cid = uuid.uuid4().hex[:16]
    logo_url = ""
    logo_file = parsed.get("logo")
    if logo_file and getattr(logo_file, "filename", None):
        try:
            saved = _save_kontakt_logo_file(logo_file, cid)
            if saved:
                logo_url = saved
        except Exception as exc:
            return jsonify({"success": False, "error": f"Logo: {exc}"}), 400

    item = {"id": cid, "logo": logo_url, **fields}
    data = load_kontakty_data()
    contacts = data.get("contacts") if isinstance(data.get("contacts"), list) else []
    contacts.append(item)
    data["contacts"] = contacts
    try:
        save_kontakty_data(data)
        log_event("kontakty_create", f"id={cid}")
        return jsonify({"success": True, "contact": item}), 201
    except Exception as exc:
        _delete_kontakt_logo_file(logo_url)
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route('/api/kontakty/<contact_id>', methods=['PUT', 'DELETE'])
@login_required
def api_kontakty_one(contact_id):
    cid = (contact_id or "").strip()[:32]
    if not cid:
        return jsonify({"success": False, "error": "Brak id"}), 400

    data = load_kontakty_data()
    contacts = data.get("contacts") if isinstance(data.get("contacts"), list) else []
    idx = next((i for i, c in enumerate(contacts) if isinstance(c, dict) and c.get("id") == cid), None)
    if idx is None:
        return jsonify({"success": False, "error": "Nie znaleziono kontaktu"}), 404

    if request.method == 'DELETE':
        old_logo = contacts[idx].get("logo") or ""
        try:
            del contacts[idx]
            data["contacts"] = contacts
            save_kontakty_data(data)
            _delete_kontakt_logo_file(old_logo)
            log_event("kontakty_delete", f"id={cid}")
            return jsonify({"success": True})
        except Exception as exc:
            return jsonify({"success": False, "error": str(exc)}), 500

    parsed = _kontakty_parse_request()
    fields = {k: parsed[k] for k in ("full_name", "company", "phone", "email")}
    if not any(fields.values()):
        return jsonify({"success": False, "error": "Uzupełnij co najmniej jedno pole."}), 400

    cur = dict(contacts[idx])
    new_logo = cur.get("logo") or ""

    if parsed.get("remove_logo"):
        _delete_kontakt_logo_file(new_logo)
        new_logo = ""

    logo_file = parsed.get("logo")
    if logo_file and getattr(logo_file, "filename", None):
        _delete_kontakt_logo_file(new_logo)
        try:
            saved = _save_kontakt_logo_file(logo_file, cid)
            new_logo = saved or ""
        except Exception as exc:
            return jsonify({"success": False, "error": f"Logo: {exc}"}), 400

    item = {"id": cid, "logo": new_logo, **fields}
    contacts[idx] = item
    data["contacts"] = contacts
    try:
        save_kontakty_data(data)
        log_event("kontakty_update", f"id={cid}")
        return jsonify({"success": True, "contact": item})
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500


@app.route('/dziennik')
@login_required
def dziennik():
    events = fetch_event_log()
    today = datetime.date.today().isoformat()
    yesterday = (datetime.date.today() - datetime.timedelta(days=1)).isoformat()
    return render_template(
        'dziennik.html',
        username=session.get('username', 'Użytkownik'),
        events=events,
        today_date=today,
        yesterday_date=yesterday,
    )

@app.route('/api/event-log', methods=['POST'])
@login_required
def event_log_api():
    payload = request.get_json(silent=True) or {}
    action = (payload.get("action") or "").strip()
    details = (payload.get("details") or "").strip()
    if not action:
        return jsonify({"success": False, "error": "Wymagana akcja"}), 400
    log_event(action, details)
    return jsonify({"success": True})

@app.route('/api/event-log/export.csv')
@login_required
def event_log_export():
    date_from = (request.args.get("date_from") or "").strip()
    date_to = (request.args.get("date_to") or "").strip()
    username = (request.args.get("user") or "").strip()
    action = (request.args.get("action") or "").strip()
    ip = (request.args.get("ip") or "").strip()

    events = fetch_event_log_filtered(
        date_from=date_from or None,
        date_to=date_to or None,
        username=username or None,
        action=action or None,
        ip=ip or None,
        limit=5000,
    )

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["created_at", "username", "action", "details", "ip"])
    for ev in events:
        writer.writerow([
            ev.get("created_at") or "",
            ev.get("username") or "",
            ev.get("action") or "",
            ev.get("details") or "",
            ev.get("ip") or "",
        ])

    resp = Response(output.getvalue(), mimetype="text/csv")
    resp.headers["Content-Disposition"] = "attachment; filename=dziennik.csv"
    return resp

# ========== KONWERTER CSV API ==========


@app.route("/api/converter/sessions", methods=["GET"])
@login_required
def converter_sessions_list_api():
    """Zwraca listę sesji konwertera dla zalogowanego użytkownika"""
    user_id = session.get("user_id")
    sessions = list_converter_sessions_for_user(user_id)
    return jsonify({"success": True, "sessions": sessions})


@app.route("/api/converter/upload", methods=["POST"])
@login_required
def converter_upload_api():
    """Upload pliku CSV i zapis jako nowa sesja konwertera"""
    if "file" not in request.files:
        return jsonify({"success": False, "error": "Brak pliku"}), 400
    file = request.files["file"]
    if not file or file.filename == "":
        return jsonify({"success": False, "error": "Brak pliku"}), 400

    try:
        raw_bytes = file.read()
        # Rozsądny limit bezpieczeństwa: ~20 MB
        if len(raw_bytes) > 20 * 1024 * 1024:
            return jsonify({"success": False, "error": "Plik jest za duży (max 20 MB)"}), 400
        csv_text = raw_bytes.decode("utf-8", errors="replace")
    except Exception:
        return jsonify({"success": False, "error": "Nie udało się odczytać pliku"}), 400

    label = (request.form.get("label") or "").strip()
    source = (request.form.get("source") or "").strip() or None
    name = label or file.filename or "Import CSV"

    user_id = session.get("user_id")
    session_id = create_converter_session(
        user_id=user_id,
        name=name,
        original_filename=file.filename,
        csv_text=csv_text,
        source=source,
    )
    log_event("converter_upload", f"session_id={session_id} name={name} filename={file.filename}")
    return jsonify({"success": True, "session_id": session_id})


@app.route("/api/converter/sessions/<session_id>/csv", methods=["GET"])
@login_required
def converter_session_csv_api(session_id):
    """Zwraca surowy CSV danej sesji (tylko dla właściciela)"""
    user_id = session.get("user_id")
    session_row = get_converter_session_csv(user_id, session_id)
    if not session_row:
        return jsonify({"success": False, "error": "Sesja nie istnieje lub brak uprawnień"}), 404
    csv_text = session_row["csv_text"]
    resp = Response(csv_text, mimetype="text/plain; charset=utf-8")
    # Dla wygody frontendu
    resp.headers["X-Converter-Name"] = session_row.get("name") or ""
    resp.headers["X-Converter-Filename"] = session_row.get("original_filename") or ""
    return resp


@app.route("/api/converter/sessions/<session_id>", methods=["DELETE"])
@login_required
def converter_session_delete_api(session_id):
    """Usuwa sesję konwertera (tylko właściciel)"""
    user_id = session.get("user_id")
    ok = delete_converter_session(user_id, session_id)
    if not ok:
        return jsonify({"success": False, "error": "Sesja nie istnieje lub brak uprawnień"}), 404
    log_event("converter_delete", f"session_id={session_id}")
    return jsonify({"success": True})


# ========== USER MANAGEMENT API ==========

@app.route('/api/users', methods=['GET', 'POST'])
@role_required('admin')
def users_api():
    """API do zarządzania użytkownikami (tylko admin)"""
    if request.method == 'GET':
        users = get_all_users()
        # Nie zwracaj password_hash
        for user in users:
            user.pop('password_hash', None)
        return jsonify({"success": True, "users": users})
    
    # POST - dodaj nowego użytkownika
    payload = request.get_json(silent=True) or {}
    username = (payload.get("username") or "").strip()
    password = payload.get("password", "")
    role = (payload.get("role") or "user").strip()
    
    if not username or not password:
        return jsonify({"success": False, "error": "Wymagane: username i password"}), 400

    if not validate_password_strength(password):
        return jsonify({"success": False, "error": "Hasło: min 9 znaków, 1 duża litera, 1 cyfra, 1 znak specjalny."}), 400
    
    if role not in ['admin', 'operator', 'user']:
        return jsonify({"success": False, "error": "Nieprawidłowa rola"}), 400
    
    user_id = create_user(username, password, role)
    if not user_id:
        return jsonify({"success": False, "error": "Użytkownik już istnieje"}), 400
    
    log_event("user_create", f"id={user_id} username={username} role={role}")
    return jsonify({"success": True, "user_id": user_id}), 201

@app.route('/api/users/<user_id>/password', methods=['PUT'])
@login_required
def user_password_api(user_id):
    """API do zmiany hasła użytkownika"""
    payload = request.get_json(silent=True) or {}
    old_password = payload.get("old_password", "")
    new_password = payload.get("new_password", "")
    is_admin_change = payload.get("is_admin_change", False)
    
    current_user_id = session.get('user_id')
    current_role = session.get('role', 'user')
    
    # Sprawdź czy użytkownik zmienia swoje własne hasło lub jest adminem
    if user_id != current_user_id and current_role != 'admin':
        return jsonify({"success": False, "error": "Brak uprawnień"}), 403
    
    if not new_password:
        return jsonify({"success": False, "error": "Wymagane nowe hasło"}), 400

    if not validate_password_strength(new_password):
        return jsonify({"success": False, "error": "Hasło: min 9 znaków, 1 duża litera, 1 cyfra, 1 znak specjalny."}), 400
    
    # Jeśli nie jest to zmiana przez admina, wymagaj starego hasła
    if not is_admin_change and user_id == current_user_id:
        if not old_password:
            return jsonify({"success": False, "error": "Wymagane stare hasło"}), 400
        
        user = get_user_by_id(user_id)
        if not user or not verify_password(user['password_hash'], old_password):
            return jsonify({"success": False, "error": "Nieprawidłowe stare hasło"}), 400
    
    updated = update_user_password(user_id, new_password)
    if not updated:
        return jsonify({"success": False, "error": "Hasło: min 9 znaków, 1 duża litera, 1 cyfra, 1 znak specjalny."}), 400
    log_event("user_password_change", f"target_id={user_id} admin_change={bool(is_admin_change)}")
    return jsonify({"success": True})

@app.route('/api/users/<user_id>/toggle', methods=['PUT'])
@role_required('admin')
def user_toggle_api(user_id):
    """API do przełączania statusu aktywnego użytkownika (tylko admin)"""
    result = toggle_user_active(user_id)
    if not result:
        return jsonify({"success": False, "error": "Użytkownik nie istnieje"}), 404
    log_event("user_toggle", f"target_id={user_id}")
    return jsonify({"success": True})

@app.route('/api/users/<user_id>/username', methods=['PUT'])
@role_required('admin')
def user_username_api(user_id):
    """API do zmiany nazwy użytkownika (tylko admin)"""
    payload = request.get_json(silent=True) or {}
    new_username = (payload.get("username") or "").strip()
    if not new_username:
        return jsonify({"success": False, "error": "Wymagana nazwa użytkownika"}), 400
    if len(new_username) < 1:
        return jsonify({"success": False, "error": "Nieprawidłowa nazwa użytkownika"}), 400
    current_user = get_user_by_id(user_id)
    if not current_user:
        return jsonify({"success": False, "error": "Użytkownik nie istnieje"}), 404
    result = update_user_username(user_id, new_username)
    if not result:
        return jsonify({"success": False, "error": "Nazwa użytkownika już istnieje"}), 400
    if session.get('user_id') == user_id:
        session['username'] = new_username
    log_event("user_rename", f"target_id={user_id} old={current_user.get('username')} new={new_username}")
    return jsonify({"success": True})

@app.route('/api/users/<user_id>/role', methods=['PUT'])
@role_required('admin')
def user_role_api(user_id):
    """API do zmiany roli użytkownika (tylko admin)"""
    payload = request.get_json(silent=True) or {}
    role = (payload.get("role") or "").strip()
    if role not in ('admin', 'operator', 'user'):
        return jsonify({"success": False, "error": "Nieprawidłowa rola"}), 400
    if not get_user_by_id(user_id):
        return jsonify({"success": False, "error": "Użytkownik nie istnieje"}), 404
    update_user_role(user_id, role)
    if session.get('user_id') == user_id:
        session['role'] = role
    log_event("user_role_change", f"target_id={user_id} new_role={role}")
    return jsonify({"success": True})

@app.route('/api/users/<user_id>', methods=['DELETE'])
@role_required('admin')
def user_delete_api(user_id):
    """API do usuwania użytkownika (tylko admin)"""
    current_user_id = session.get('user_id')
    if user_id == current_user_id:
        return jsonify({"success": False, "error": "Nie możesz usunąć własnego konta"}), 400
    
    target = get_user_by_id(user_id)
    if not target:
        return jsonify({"success": False, "error": "Użytkownik nie istnieje"}), 404
    with closing(get_reklamacje_db()) as conn:
        delete_sql = "DELETE FROM users WHERE id = %s" if using_postgres() else "DELETE FROM users WHERE id = ?"
        conn.execute(delete_sql, (user_id,))
        conn.commit()
    log_event("user_delete", f"target_id={user_id} username={target.get('username')}")
    return jsonify({"success": True})

# Jawna obsługa plików statycznych (backup)
@app.route('/static/<path:filename>')
def static_files(filename):
    return send_from_directory(app.static_folder, filename)

# Dla produkcji (Docker/WSGI)
application = app

if __name__ == '__main__':
    # Dla lokalnego uruchomienia
    app.run(debug=True, port=5000, host='0.0.0.0')
